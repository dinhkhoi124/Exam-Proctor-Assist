from __future__ import annotations

from datetime import date, datetime, time, timedelta
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models.chat_log import ChatLog
from app.models.chat_session import ChatSession  # noqa: F401 - registers relationship mapper
from app.models.chat_topic import ChatTopic
from app.models.feedback_log import FeedbackLog
from app.models.user import User


MAX_DETAIL_ROWS = 50_000
GROUP_FORMATS = {
    "day": "%d/%m/%Y",
    "week": "Tuan %W - %Y",
    "month": "%m/%Y",
    "year": "%Y",
}


def _date_bounds(start_date: date, end_date: date) -> tuple[datetime, datetime]:
    return (
        datetime.combine(start_date, time.min),
        datetime.combine(end_date + timedelta(days=1), time.min),
    )


def _apply_chat_filters(query, start_at: datetime, end_at: datetime, topic_id: int | None):
    query = query.filter(ChatLog.created_at >= start_at, ChatLog.created_at < end_at)
    if topic_id is not None:
        query = query.filter(ChatLog.topic_id == topic_id)
    return query


def collect_report_data(
    db: Session,
    start_date: date,
    end_date: date,
    group_by: str,
    topic_id: int | None = None,
    include_details: bool = False,
) -> dict[str, Any]:
    start_at, end_at = _date_bounds(start_date, end_date)
    period = func.date_trunc(group_by, ChatLog.created_at)

    summary_row = _apply_chat_filters(
        db.query(
            func.count(ChatLog.id).label("questions"),
            func.count(func.distinct(ChatLog.user_id)).label("active_users"),
            func.avg(ChatLog.latency_ms).label("avg_latency_ms"),
        ),
        start_at,
        end_at,
        topic_id,
    ).one()

    feedback_query = (
        db.query(FeedbackLog.rating, func.count(FeedbackLog.id))
        .join(ChatLog, FeedbackLog.chat_id == ChatLog.id)
        .filter(FeedbackLog.is_deleted.is_(False))
    )
    feedback_query = _apply_chat_filters(feedback_query, start_at, end_at, topic_id)
    feedback_rows = feedback_query.group_by(FeedbackLog.rating).all()
    feedback = {"like": 0, "dislike": 0}
    for rating, count in feedback_rows:
        if rating in feedback:
            feedback[rating] = count

    timeline_query = _apply_chat_filters(
        db.query(
            period.label("period"),
            func.count(ChatLog.id).label("questions"),
            func.count(func.distinct(ChatLog.user_id)).label("users"),
            func.avg(ChatLog.latency_ms).label("avg_latency_ms"),
        ),
        start_at,
        end_at,
        topic_id,
    )
    timeline_rows = timeline_query.group_by(period).order_by(period).all()

    topic_query = _apply_chat_filters(
        db.query(
            func.coalesce(ChatTopic.name, "General Guidance").label("topic"),
            func.count(ChatLog.id).label("questions"),
        ).outerjoin(ChatTopic, ChatLog.topic_id == ChatTopic.id),
        start_at,
        end_at,
        topic_id,
    )
    topic_rows = topic_query.group_by(ChatTopic.name).order_by(func.count(ChatLog.id).desc()).all()

    user_query = _apply_chat_filters(
        db.query(
            User.username,
            User.email,
            func.count(ChatLog.id).label("questions"),
        ).outerjoin(User, ChatLog.user_id == User.id),
        start_at,
        end_at,
        topic_id,
    )
    user_rows = (
        user_query.group_by(User.id, User.username, User.email)
        .order_by(func.count(ChatLog.id).desc())
        .limit(20)
        .all()
    )

    total_feedback = feedback["like"] + feedback["dislike"]
    result: dict[str, Any] = {
        "filters": {
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "group_by": group_by,
            "topic_id": topic_id,
        },
        "summary": {
            "total_questions": summary_row.questions or 0,
            "active_users": summary_row.active_users or 0,
            "avg_latency_ms": round(float(summary_row.avg_latency_ms or 0)),
            "likes": feedback["like"],
            "dislikes": feedback["dislike"],
            "satisfaction_rate": round(feedback["like"] * 100 / total_feedback, 1)
            if total_feedback
            else 0,
        },
        "timeline": [
            {
                "period": row.period.isoformat(),
                "label": row.period.strftime(GROUP_FORMATS[group_by]),
                "questions": row.questions,
                "users": row.users,
                "avg_latency_ms": round(float(row.avg_latency_ms or 0)),
            }
            for row in timeline_rows
        ],
        "topics": [{"topic": row.topic, "questions": row.questions} for row in topic_rows],
        "top_users": [
            {
                "username": row.username or "Unknown",
                "email": row.email or "",
                "questions": row.questions,
            }
            for row in user_rows
        ],
    }

    if include_details:
        detail_query = _apply_chat_filters(
            db.query(ChatLog, User, ChatTopic)
            .outerjoin(User, ChatLog.user_id == User.id)
            .outerjoin(ChatTopic, ChatLog.topic_id == ChatTopic.id),
            start_at,
            end_at,
            topic_id,
        )
        detail_rows = detail_query.order_by(ChatLog.created_at.desc()).limit(MAX_DETAIL_ROWS + 1).all()
        result["details_truncated"] = len(detail_rows) > MAX_DETAIL_ROWS
        result["details"] = [
            {
                "created_at": chat.created_at,
                "username": user.username if user else "Unknown",
                "email": user.email if user else "",
                "topic": topic.name if topic else "General Guidance",
                "question": chat.question,
                "answer": chat.answer or "",
                "latency_ms": chat.latency_ms or 0,
                "status": chat.status or "",
            }
            for chat, user, topic in detail_rows[:MAX_DETAIL_ROWS]
        ]

    return result


def build_excel_report(data: dict[str, Any]) -> BytesIO:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Tong quan"

    filters = data["filters"]
    summary = data["summary"]
    overview.append(["BAO CAO HOAT DONG CHATBOT"])
    overview.append(["Tu ngay", filters["start_date"]])
    overview.append(["Den ngay", filters["end_date"]])
    overview.append([])
    overview.append(["Chi so", "Gia tri"])
    overview.append(["Tong cau hoi", summary["total_questions"]])
    overview.append(["Nguoi dung hoat dong", summary["active_users"]])
    overview.append(["Thoi gian phan hoi trung binh (ms)", summary["avg_latency_ms"]])
    overview.append(["Luot hai long", summary["likes"]])
    overview.append(["Luot chua hai long", summary["dislikes"]])
    overview.append(["Ty le hai long (%)", summary["satisfaction_rate"]])

    _append_table(
        workbook,
        "Theo thoi gian",
        ["Thoi gian", "So cau hoi", "Nguoi dung", "Phan hoi TB (ms)"],
        [
            [row["label"], row["questions"], row["users"], row["avg_latency_ms"]]
            for row in data["timeline"]
        ],
    )
    _append_table(
        workbook,
        "Chu de",
        ["Chu de", "So cau hoi"],
        [[row["topic"], row["questions"]] for row in data["topics"]],
    )
    _append_table(
        workbook,
        "Nguoi dung",
        ["Ten dang nhap", "Email", "So cau hoi"],
        [[row["username"], row["email"], row["questions"]] for row in data["top_users"]],
    )
    _append_table(
        workbook,
        "Chi tiet",
        ["Thoi gian", "Nguoi dung", "Email", "Chu de", "Cau hoi", "Cau tra loi", "Latency (ms)", "Trang thai"],
        [
            [
                row["created_at"],
                row["username"],
                row["email"],
                row["topic"],
                row["question"],
                row["answer"],
                row["latency_ms"],
                row["status"],
            ]
            for row in data.get("details", [])
        ],
    )

    if data.get("details_truncated"):
        overview.append([])
        overview.append([f"Chi tiet duoc gioi han o {MAX_DETAIL_ROWS:,} dong."])

    for sheet in workbook.worksheets:
        _style_sheet(sheet)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _append_table(workbook: Workbook, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)


def _style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="F97316")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in range(1, sheet.max_column + 1):
        max_length = min(
            max((len(str(sheet.cell(row, column).value or "")) for row in range(1, sheet.max_row + 1)), default=10)
            + 2,
            60,
        )
        sheet.column_dimensions[get_column_letter(column)].width = max_length
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_pdf_report(data: dict[str, Any]) -> BytesIO:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages

    plt.rcParams["font.family"] = "DejaVu Sans"
    output = BytesIO()
    summary = data["summary"]
    filters = data["filters"]

    with PdfPages(output) as pdf:
        figure = plt.figure(figsize=(11.69, 8.27))
        grid = figure.add_gridspec(2, 2, height_ratios=[1, 2.5])
        figure.suptitle("Báo cáo hoạt động Chatbot", fontsize=20, fontweight="bold", color="#c2410c")
        figure.text(
            0.5,
            0.91,
            f"Từ {filters['start_date']} đến {filters['end_date']}",
            ha="center",
            fontsize=10,
            color="#475569",
        )

        kpi_axis = figure.add_subplot(grid[0, :])
        kpi_axis.axis("off")
        kpi_axis.table(
            cellText=[[
                f"{summary['total_questions']:,}",
                f"{summary['active_users']:,}",
                f"{summary['avg_latency_ms']:,} ms",
                f"{summary['satisfaction_rate']}%",
            ]],
            colLabels=["Tổng câu hỏi", "Người dùng hoạt động", "Phản hồi trung bình", "Tỷ lệ hài lòng"],
            loc="center",
            cellLoc="center",
        ).scale(1, 2)

        timeline_axis = figure.add_subplot(grid[1, 0])
        labels = [row["label"] for row in data["timeline"]]
        timeline_axis.plot(labels, [row["questions"] for row in data["timeline"]], marker="o", label="Câu hỏi")
        timeline_axis.plot(labels, [row["users"] for row in data["timeline"]], marker="o", label="Người dùng")
        timeline_axis.set_title("Hoạt động theo thời gian")
        timeline_axis.tick_params(axis="x", rotation=45, labelsize=7)
        timeline_axis.grid(alpha=0.25)
        timeline_axis.legend()

        topic_axis = figure.add_subplot(grid[1, 1])
        top_topics = data["topics"][:10]
        if top_topics:
            topic_axis.barh(
                [row["topic"] for row in reversed(top_topics)],
                [row["questions"] for row in reversed(top_topics)],
                color="#f97316",
            )
        topic_axis.set_title("Top chủ đề")
        topic_axis.grid(axis="x", alpha=0.25)
        figure.tight_layout(rect=[0.03, 0.03, 0.97, 0.89])
        pdf.savefig(figure)
        plt.close(figure)

        figure, axes = plt.subplots(1, 2, figsize=(11.69, 8.27))
        figure.suptitle("Phản hồi và người dùng nổi bật", fontsize=18, fontweight="bold", color="#c2410c")

        feedback_values = [summary["likes"], summary["dislikes"]]
        if sum(feedback_values):
            axes[0].pie(
                feedback_values,
                labels=["Hài lòng", "Chưa hài lòng"],
                autopct="%1.1f%%",
                colors=["#22c55e", "#ef4444"],
            )
        else:
            axes[0].text(0.5, 0.5, "Chưa có phản hồi", ha="center", va="center")
        axes[0].set_title("Mức độ hài lòng")

        axes[1].axis("off")
        top_users = data["top_users"][:10]
        if top_users:
            axes[1].table(
                cellText=[[row["username"], row["questions"]] for row in top_users],
                colLabels=["Người dùng", "Số câu hỏi"],
                loc="center",
                cellLoc="left",
            ).scale(1, 1.5)
        else:
            axes[1].text(0.5, 0.5, "Chưa có dữ liệu người dùng", ha="center", va="center")
        axes[1].set_title("Top người dùng")
        figure.tight_layout(rect=[0.03, 0.03, 0.97, 0.92])
        pdf.savefig(figure)
        plt.close(figure)

    output.seek(0)
    return output
