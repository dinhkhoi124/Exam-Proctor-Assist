import os
import csv
import json
import hashlib
from pathlib import Path

# Paths
REPO_ROOT = Path("e:/merged_partition_content/Khoi_Project/FPT-Assistant-v3")
BENCHMARK_DIR = REPO_ROOT / "dataset_benchmark"
OUTPUT_DIR = BENCHMARK_DIR / "benchmark_outputs"
ANNOTATION_DIR = BENCHMARK_DIR / "annotations"

# Files
manifest_path = BENCHMARK_DIR / "manifest.csv"
raw_transcripts_path = OUTPUT_DIR / "raw_transcripts.jsonl"
corrected_transcripts_path = OUTPUT_DIR / "corrected_transcripts.jsonl"
transcript_per_sample_path = OUTPUT_DIR / "transcript_per_sample.json"
retrieval_proxy_per_sample_path = OUTPUT_DIR / "retrieval_proxy_per_sample.json"
run_manifest_path = OUTPUT_DIR / "run_manifest.json"

def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()

def load_jsonl(path: Path) -> dict:
    if not path.exists():
        return {}
    data = {}
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            if line.strip():
                row = json.loads(line)
                data[int(row["audio_id"])] = row
    return data

def load_json(path: Path):
    if not path.exists():
        return None
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)

def load_csv(path: Path) -> list:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))

def main():
    print("--- RUNNING AUDIT SCRIPT ---")
    
    # 1. Load data
    manifest_rows = load_csv(manifest_path)
    # Filter test split
    test_manifest = [r for r in manifest_rows if r["split"] == "test" and r["eligibility_status"] == "eligible"]
    test_ids = {int(r["audio_id"]) for r in test_manifest}
    speaker_map = {int(r["audio_id"]): r["speaker"] for r in test_manifest}
    
    raw_data = load_jsonl(raw_transcripts_path)
    corrected_data = load_jsonl(corrected_transcripts_path)
    transcripts_per_sample = load_json(transcript_per_sample_path)
    retrieval_proxy_per_sample = load_json(retrieval_proxy_per_sample_path)
    
    # Check alignment
    print(f"Loaded {len(test_manifest)} test manifest rows.")
    print(f"Loaded {len(transcripts_per_sample)} transcript_per_sample rows.")
    print(f"Loaded {len(retrieval_proxy_per_sample)} retrieval_proxy_per_sample rows.")
    
    # Map transcripts per sample
    transcripts_map = {int(item["audio_id"]): item for item in transcripts_per_sample}
    proxy_map = {int(item["audio_id"]): item for item in retrieval_proxy_per_sample}
    
    # 2. Build full per-sample CSV table
    # Columns:
    # audio_id, speaker, raw_transcript, corrected_transcript, reference_transcript,
    # raw_wer, corrected_wer, delta_wer, raw_cer, corrected_cer, delta_cer,
    # is_changed, category, correction_status
    persample_rows = []
    
    for aid in sorted(test_ids):
        manifest_row = speaker_map[aid]
        raw_row = raw_data.get(aid, {})
        corrected_row = corrected_data.get(aid, {})
        ts_row = transcripts_map.get(aid, {})
        
        raw_text = ts_row.get("raw_transcript", "")
        corrected_text = ts_row.get("corrected_transcript", "")
        reference_text = ts_row.get("reference_transcript", "")
        
        raw_wer = ts_row.get("baseline_wer", 0.0)
        corrected_wer = ts_row.get("proposed_wer", 0.0)
        delta_wer = raw_wer - corrected_wer
        
        raw_cer = ts_row.get("baseline_cer", 0.0)
        corrected_cer = ts_row.get("proposed_cer", 0.0)
        delta_cer = raw_cer - corrected_cer
        
        is_changed = raw_text != corrected_text
        category = ts_row.get("outcome", "unchanged")
        correction_status = corrected_row.get("status", "success")
        
        persample_rows.append({
            "audio_id": aid,
            "speaker": speaker_map[aid],
            "raw_transcript": raw_text,
            "corrected_transcript": corrected_text,
            "reference_transcript": reference_text,
            "raw_wer": f"{raw_wer:.6f}",
            "corrected_wer": f"{corrected_wer:.6f}",
            "delta_wer": f"{delta_wer:.6f}",
            "raw_cer": f"{raw_cer:.6f}",
            "corrected_cer": f"{corrected_cer:.6f}",
            "delta_cer": f"{delta_cer:.6f}",
            "is_changed": str(is_changed),
            "category": category,
            "correction_status": correction_status
        })
        
    # Write test_persample_audit.csv
    persample_headers = [
        "audio_id", "speaker", "raw_transcript", "corrected_transcript", "reference_transcript",
        "raw_wer", "corrected_wer", "delta_wer", "raw_cer", "corrected_cer", "delta_cer",
        "is_changed", "category", "correction_status"
    ]
    persample_csv_path = REPO_ROOT / "test_persample_audit.csv"
    with persample_csv_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=persample_headers)
        writer.writeheader()
        writer.writerows(persample_rows)
    print(f"Wrote {len(persample_rows)} rows to test_persample_audit.csv")

    # 3. Verification of aggregated WER
    # We will sum errors and reference words over N=104
    total_baseline_word_errors = 0
    total_proposed_word_errors = 0
    total_reference_words = 0
    
    for aid in test_ids:
        ts_row = transcripts_map.get(aid, {})
        total_baseline_word_errors += ts_row.get("baseline_word_errors", 0)
        total_proposed_word_errors += ts_row.get("proposed_word_errors", 0)
        total_reference_words += ts_row.get("reference_words", 0)
        
    baseline_agg_wer = total_baseline_word_errors / max(1, total_reference_words)
    proposed_agg_wer = total_proposed_word_errors / max(1, total_reference_words)
    
    print("\n--- AGGREGATE WER VERIFICATION ---")
    print(f"Total reference words: {total_reference_words}")
    print(f"Baseline (raw):")
    print(f"  Total edit distance (word errors): {total_baseline_word_errors}")
    print(f"  WER: {baseline_agg_wer * 100:.6f}%")
    print(f"Proposed (corrected):")
    print(f"  Total edit distance (word errors): {total_proposed_word_errors}")
    print(f"  WER: {proposed_agg_wer * 100:.6f}%")
    
    # 4. Extract degraded and improved rows
    degraded_improved_rows = [r for r in persample_rows if r["category"] in ("degraded", "improved")]
    degraded_improved_csv_path = REPO_ROOT / "test_degraded_improved_detail.csv"
    with degraded_improved_csv_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=persample_headers)
        writer.writeheader()
        writer.writerows(degraded_improved_rows)
    print(f"\nWrote {len(degraded_improved_rows)} rows to test_degraded_improved_detail.csv")
    
    # 5. Extract unchanged with errors (category == "unchanged" and raw_wer > 0)
    unchanged_with_errors_rows = [r for r in persample_rows if r["category"] == "unchanged" and float(r["raw_wer"]) > 0.0]
    unchanged_with_errors_csv_path = REPO_ROOT / "test_unchanged_with_errors.csv"
    with unchanged_with_errors_csv_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=persample_headers)
        writer.writeheader()
        writer.writerows(unchanged_with_errors_rows)
    print(f"Wrote {len(unchanged_with_errors_rows)} rows to test_unchanged_with_errors.csv")
    
    # 6. Retrieval proxy per-sample table
    # Columns: audio_id, baseline_jaccard_at_5, baseline_overlap_recall_at_5, proposed_jaccard_at_5, proposed_overlap_recall_at_5
    proxy_rows_out = []
    for aid in sorted(test_ids):
        p_row = proxy_map.get(aid, {})
        baseline_jac = p_row.get("baseline", {}).get("jaccard_at_5", 0.0)
        baseline_rec = p_row.get("baseline", {}).get("overlap_recall_at_5", 0.0)
        proposed_jac = p_row.get("proposed", {}).get("jaccard_at_5", 0.0)
        proposed_rec = p_row.get("proposed", {}).get("overlap_recall_at_5", 0.0)
        
        proxy_rows_out.append({
            "audio_id": aid,
            "baseline_jaccard_at_5": f"{baseline_jac:.6f}",
            "baseline_overlap_recall_at_5": f"{baseline_rec:.6f}",
            "proposed_jaccard_at_5": f"{proposed_jac:.6f}",
            "proposed_overlap_recall_at_5": f"{proposed_rec:.6f}"
        })
    proxy_headers = [
        "audio_id", "baseline_jaccard_at_5", "baseline_overlap_recall_at_5",
        "proposed_jaccard_at_5", "proposed_overlap_recall_at_5"
    ]
    proxy_csv_path = REPO_ROOT / "test_retrieval_proxy_persample.csv"
    with proxy_csv_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=proxy_headers)
        writer.writeheader()
        writer.writerows(proxy_rows_out)
    print(f"Wrote {len(proxy_rows_out)} rows to test_retrieval_proxy_persample.csv")

    # 7. Check human/gold annotation status from live xlsx files if openpyxl is installed
    print("\n--- LIVE ANNOTATION STATUS ---")
    try:
        import openpyxl
        
        def count_gold_xlsx(path: Path):
            if not path.exists():
                return None
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            sheet = wb["Gold"]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return 0
            headers = [str(v) if v is not None else "" for v in rows[0]]
            positions = {header: idx for idx, header in enumerate(headers)}
            
            source_col = positions.get("gold_source")
            pages_col = positions.get("gold_pages")
            
            count = 0
            for r in rows[1:]:
                # Check if audio_id is not null
                if r[positions.get("audio_id")] is None:
                    continue
                src = r[source_col] if source_col is not None else None
                pgs = r[pages_col] if pages_col is not None else None
                if (src is not None and str(src).strip()) or (pgs is not None and str(pgs).strip()):
                    count += 1
            return count

        def count_graded_xlsx(path: Path):
            if not path.exists():
                return None
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            sheet = wb["Blind grading"]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return 0
            headers = [str(v) if v is not None else "" for v in rows[0]]
            positions = {header: idx for idx, header in enumerate(headers)}
            
            # Graded if any score column is filled
            score_cols = [
                positions.get(h) for h in headers 
                if any(x in h for x in ("correctness", "faithfulness", "completeness", "citation", "source_page_correct", "task_success"))
            ]
            score_cols = [c for c in score_cols if c is not None]
            
            count = 0
            for r in rows[1:]:
                if r[positions.get("audio_id")] is None:
                    continue
                is_graded = False
                for c in score_cols:
                    if r[c] is not None and str(r[c]).strip() != "":
                        is_graded = True
                        break
                if is_graded:
                    count += 1
            return count

        gold_a = count_gold_xlsx(ANNOTATION_DIR / "gold_rater_A.xlsx")
        gold_b = count_gold_xlsx(ANNOTATION_DIR / "gold_rater_B.xlsx")
        
        # Check adjudicated
        adjudicated_path = ANNOTATION_DIR / "gold_adjudicated.xlsx"
        adjudicated_count = None
        if adjudicated_path.exists():
            wb = openpyxl.load_workbook(adjudicated_path, read_only=True, data_only=True)
            sheet = wb["Adjudication"]
            rows = list(sheet.iter_rows(values_only=True))
            if rows:
                headers = [str(v) if v is not None else "" for v in rows[0]]
                positions = {header: idx for idx, header in enumerate(headers)}
                status_col = positions.get("agreement_status")
                
                adjudicated_count = 0
                for r in rows[1:]:
                    if r[positions.get("audio_id")] is None:
                        continue
                    status = str(r[status_col]).strip().casefold() if status_col is not None and r[status_col] is not None else ""
                    if status in ("agreed", "adjudicated"):
                        adjudicated_count += 1
        
        answers_a = count_graded_xlsx(ANNOTATION_DIR / "answers_rater_A.xlsx")
        answers_b = count_graded_xlsx(ANNOTATION_DIR / "answers_rater_B.xlsx")
        
        print(f"Rater A Gold source/page completed: {gold_a}/60")
        print(f"Rater B Gold source/page completed: {gold_b}/60")
        if adjudicated_count is not None:
            print(f"Gold Adjudicated completed: {adjudicated_count}/60")
        else:
            print("Gold Adjudicated file (gold_adjudicated.xlsx) does not exist.")
        print(f"Rater A answers graded: {answers_a}/60")
        print(f"Rater B answers graded: {answers_b}/60")
        
    except Exception as e:
        print(f"Error reading xlsx files: {e}")

    # 8. Prompt SHA-256 Check
    print("\n--- PROMPT SHA-256 CHECK ---")
    current_prompt_path = REPO_ROOT / "backend/app/prompts/asr_correction.py"
    if current_prompt_path.exists():
        curr_hash = sha256_file(current_prompt_path)
        locked_hash = "9ea1a085b9887dba7781a7fb620063e70325891edadad6dbcc94e798be7bde1d"
        print(f"Current Prompt SHA-256: {curr_hash}")
        print(f"Locked  Prompt SHA-256: {locked_hash}")
        if curr_hash == locked_hash:
            print("MATCH: Current prompt hash matches the locked prompt hash exactly.")
        else:
            print("MISMATCH: Current prompt hash does NOT match the locked prompt hash!")
    else:
        print(f"Prompt file not found at {current_prompt_path}")

if __name__ == "__main__":
    main()
