from __future__ import annotations

from collections import Counter, defaultdict
import csv
import json
from pathlib import Path
from statistics import mean
from typing import Iterable

try:
    from .common import atomic_write_json, load_config, read_jsonl, resolve_path
    from .metrics import (
        aggregate_transcripts,
        bootstrap_mean_difference,
        build_transcript_rows,
        proxy_retrieval_metrics,
        true_retrieval_metrics,
    )
    from .annotations import load_human_grades
except ImportError:
    from common import atomic_write_json, load_config, read_jsonl, resolve_path
    from metrics import (
        aggregate_transcripts,
        bootstrap_mean_difference,
        build_transcript_rows,
        proxy_retrieval_metrics,
        true_retrieval_metrics,
    )
    from annotations import load_human_grades


def _read_manifest(path: Path) -> list[dict]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _percentile(values: list[float], percentile: float) -> float | None:
    if not values:
        return None
    ordered = sorted(values)
    index = (len(ordered) - 1) * percentile
    lower = int(index)
    upper = min(len(ordered) - 1, lower + 1)
    fraction = index - lower
    return ordered[lower] * (1 - fraction) + ordered[upper] * fraction


def _average(rows: Iterable[dict], key: str) -> float:
    values = [float(row[key]) for row in rows]
    return mean(values) if values else 0.0


def _parse_pages(value: object) -> list[int]:
    if value is None:
        return []
    return [int(item.strip()) for item in str(value).split(";") if item.strip().isdigit()]


def _completed_human_grades(grades: dict, rubric_names: tuple[str, ...]) -> dict:
    return {
        key: scores
        for key, scores in grades.items()
        if any(scores.get(rubric) is not None for rubric in rubric_names)
    }


def load_adjudicated_gold(path: Path) -> dict[int, set[tuple[str, int]]]:
    if not path.exists():
        return {}
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError:
        return {}
    sheet = load_workbook(path, read_only=True, data_only=True).active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    positions = {str(header): index for index, header in enumerate(headers)}
    output = {}
    for values in sheet.iter_rows(min_row=2, values_only=True):
        audio_id = values[positions["audio_id"]]
        source = values[positions["gold_source"]]
        pages = values[positions["gold_pages"]]
        status = (
            str(values[positions["agreement_status"]]).strip().casefold()
            if "agreement_status" in positions
            and values[positions["agreement_status"]] is not None
            else ""
        )
        if status and status not in {"agreed", "adjudicated"}:
            continue
        if audio_id is None or not source:
            continue
        output[int(audio_id)] = {
            (str(source).strip(), page) for page in _parse_pages(pages)
        }
    return output


def _paired_test(baseline: list[float], proposed: list[float]) -> dict:
    if not baseline or len(baseline) != len(proposed):
        return {}
    try:
        from scipy.stats import rankdata, wilcoxon

        deltas = [base - prop for base, prop in zip(baseline, proposed)]
        nonzero = [delta for delta in deltas if delta != 0]
        if not nonzero:
            return {"wilcoxon_p": 1.0, "rank_biserial": 0.0}
        statistic = wilcoxon(baseline, proposed, zero_method="wilcox")
        ranks = rankdata([abs(delta) for delta in nonzero])
        positive = sum(rank for rank, delta in zip(ranks, nonzero) if delta > 0)
        negative = sum(rank for rank, delta in zip(ranks, nonzero) if delta < 0)
        return {
            "wilcoxon_p": float(statistic.pvalue),
            "rank_biserial": float((positive - negative) / (positive + negative)),
        }
    except (ImportError, ValueError):
        return {}


def _mcnemar(baseline: list[bool], proposed: list[bool]) -> dict:
    baseline_only = sum(base and not prop for base, prop in zip(baseline, proposed))
    proposed_only = sum(prop and not base for base, prop in zip(baseline, proposed))
    discordant = baseline_only + proposed_only
    try:
        from scipy.stats import binomtest

        p_value = (
            float(binomtest(min(baseline_only, proposed_only), discordant, 0.5).pvalue)
            if discordant
            else 1.0
        )
    except ImportError:
        p_value = None
    return {
        "baseline_only": baseline_only,
        "proposed_only": proposed_only,
        "p_value": p_value,
        "matched_odds_ratio": (
            proposed_only / baseline_only if baseline_only else None
        ),
    }


def evaluate(config: dict) -> dict:
    output_dir = resolve_path(config["output_dir"])
    manifest = _read_manifest(resolve_path(config["manifest"]))
    requested_split = config.get("_sample_split", "all")
    selected_manifest = [
        row for row in manifest if row["eligibility_status"] == "eligible"
    ]
    if requested_split in {"dev", "test"}:
        selected_manifest = [
            row for row in selected_manifest if row["split"] == requested_split
        ]
    limit = config.get("_limit")
    if limit is not None:
        selected_manifest = selected_manifest[: int(limit)]

    # A normal full run reports locked test metrics. Explicit dev smoke runs report
    # only the selected dev subset, so cached records from other runs cannot leak in.
    evaluation_split = "test" if requested_split == "all" else requested_split
    evaluation_manifest = [
        row
        for row in selected_manifest
        if requested_split != "all" or row["split"] == "test"
    ]
    evaluation_ids = {int(row["audio_id"]) for row in evaluation_manifest}
    raw = {int(row["audio_id"]): row for row in read_jsonl(output_dir / "raw_transcripts.jsonl")}
    corrected = {int(row["audio_id"]): row for row in read_jsonl(output_dir / "corrected_transcripts.jsonl")}
    retrieval = {int(row["audio_id"]): row for row in read_jsonl(output_dir / "retrieval_results.jsonl")}
    answers = {int(row["audio_id"]): row for row in read_jsonl(output_dir / "answer_results.jsonl")}
    judge = {int(row["audio_id"]): row for row in read_jsonl(output_dir / "judge_results.jsonl")}

    raw = {audio_id: row for audio_id, row in raw.items() if audio_id in evaluation_ids}
    corrected = {
        audio_id: row for audio_id, row in corrected.items() if audio_id in evaluation_ids
    }
    retrieval = {
        audio_id: row for audio_id, row in retrieval.items() if audio_id in evaluation_ids
    }
    answers = {
        audio_id: row for audio_id, row in answers.items() if audio_id in evaluation_ids
    }
    judge = {audio_id: row for audio_id, row in judge.items() if audio_id in evaluation_ids}

    transcript_rows = build_transcript_rows(
        evaluation_manifest, raw, corrected, split=None
    )
    transcript_summary = aggregate_transcripts(transcript_rows) if transcript_rows else {}
    if transcript_rows:
        transcript_summary["wer_bootstrap_95_ci"] = bootstrap_mean_difference(
            [row["baseline_wer"] for row in transcript_rows],
            [row["proposed_wer"] for row in transcript_rows],
            iterations=int(config.get("bootstrap_iterations", 10000)),
            seed=int(config.get("split_seed", 42)),
        )
        transcript_summary["paired_test"] = _paired_test(
            [row["baseline_wer"] for row in transcript_rows],
            [row["proposed_wer"] for row in transcript_rows],
        )

    proxy_rows = []
    for audio_id in sorted(evaluation_ids & retrieval.keys()):
        values = proxy_retrieval_metrics(retrieval[audio_id]["retrieval"])
        proxy_rows.append({"audio_id": audio_id, **values})
    proxy_summary = {
        branch: {
            "jaccard_at_5": _average(
                [row[branch] for row in proxy_rows], "jaccard_at_5"
            ),
            "overlap_recall_at_5": _average(
                [row[branch] for row in proxy_rows], "overlap_recall_at_5"
            ),
        }
        for branch in ("baseline", "proposed")
    }
    if proxy_rows:
        for metric in ("jaccard_at_5", "overlap_recall_at_5"):
            baseline_values = [row["baseline"][metric] for row in proxy_rows]
            proposed_values = [row["proposed"][metric] for row in proxy_rows]
            proxy_summary[f"{metric}_bootstrap_95_ci"] = bootstrap_mean_difference(
                baseline_values,
                proposed_values,
                iterations=int(config.get("bootstrap_iterations", 10000)),
                seed=int(config.get("split_seed", 42)),
            )
            proxy_summary[f"{metric}_paired_test"] = _paired_test(
                baseline_values, proposed_values
            )

    gold = load_adjudicated_gold(
        resolve_path(config["annotation_dir"]) / "gold_adjudicated.xlsx"
    )
    true_rows = []
    for audio_id in sorted(evaluation_ids & retrieval.keys() & gold.keys()):
        row = {"audio_id": audio_id}
        for branch in ("baseline", "proposed"):
            candidates = retrieval[audio_id]["retrieval"][branch]["candidates"]
            row[branch] = true_retrieval_metrics(candidates, gold[audio_id])
        true_rows.append(row)
    true_summary = {}
    if true_rows:
        for branch in ("baseline", "proposed"):
            true_summary[branch] = {
                key: _average([row[branch] for row in true_rows], key)
                for key in (
                    "hit_at_1",
                    "hit_at_3",
                    "hit_at_5",
                    "mrr_at_10",
                    "recall_at_5",
                    "ndcg_at_5",
                )
            }
        true_summary["mcnemar_hit_at_5"] = _mcnemar(
            [row["baseline"]["hit_at_5"] for row in true_rows],
            [row["proposed"]["hit_at_5"] for row in true_rows],
        )
        true_summary["mrr_bootstrap_95_ci"] = bootstrap_mean_difference(
            [row["baseline"]["mrr_at_10"] for row in true_rows],
            [row["proposed"]["mrr_at_10"] for row in true_rows],
            iterations=int(config.get("bootstrap_iterations", 10000)),
            seed=int(config.get("split_seed", 42)),
        )
        true_summary["mrr_paired_test"] = _paired_test(
            [row["baseline"]["mrr_at_10"] for row in true_rows],
            [row["proposed"]["mrr_at_10"] for row in true_rows],
        )

    grades, _mappings = load_human_grades(resolve_path(config["annotation_dir"]))
    rubric_names = (
        "correctness",
        "faithfulness",
        "completeness",
        "citation",
        "source_page_correct",
        "task_success",
    )
    completed_grades = _completed_human_grades(grades, rubric_names)
    human_summary = {
        "template_records": len(grades),
        "graded_records": len(completed_grades),
        "graded_audio_ids": len(
            {audio_id for (_rater, audio_id, _branch) in completed_grades}
        ),
    }
    for branch in ("baseline", "proposed"):
        human_summary[branch] = {}
        for rubric in rubric_names:
            values = [
                float(scores[rubric])
                for (_rater, item_audio_id, item_branch), scores in completed_grades.items()
                if item_audio_id in evaluation_ids
                and item_branch == branch
                and scores.get(rubric) is not None
            ]
            human_summary[branch][rubric] = mean(values) if values else None
    agreement = {}
    try:
        from sklearn.metrics import cohen_kappa_score

        for rubric in rubric_names:
            first, second = [], []
            keys = sorted(
                (audio_id, branch)
                for rater, audio_id, branch in completed_grades
                if rater == "A"
                and audio_id in evaluation_ids
                and ("B", audio_id, branch) in completed_grades
            )
            for audio_id, branch in keys:
                value_a = completed_grades[("A", audio_id, branch)].get(rubric)
                value_b = completed_grades[("B", audio_id, branch)].get(rubric)
                if value_a is not None and value_b is not None:
                    first.append(int(value_a))
                    second.append(int(value_b))
            agreement[rubric] = (
                float(cohen_kappa_score(first, second, weights="quadratic"))
                if first
                else None
            )
    except ImportError:
        pass
    human_summary["weighted_cohen_kappa"] = agreement

    judge_summary = {"samples": sum(row.get("status") == "success" for row in judge.values())}
    for branch in ("baseline", "proposed"):
        judge_summary[branch] = {}
        for rubric in ("correctness", "faithfulness", "completeness", "citation", "task_success"):
            values = [
                float(row["scores"][branch][rubric])
                for row in judge.values()
                if row.get("status") == "success"
                and branch in row.get("scores", {})
                and rubric in row["scores"][branch]
            ]
            judge_summary[branch][rubric] = mean(values) if values else None

    pricing = config.get("pricing", {})
    price_models = pricing.get("models", {})
    usage_cost = 0.0
    token_cost_by_model: dict[str, float] = defaultdict(float)
    stt_cost = 0.0
    priced = False
    token_groups = [
        (row.get("model"), row.get("prompt_tokens") or 0, row.get("completion_tokens") or 0)
        for row in corrected.values()
    ]
    for row in answers.values():
        for branch in ("baseline", "proposed"):
            token_groups.append(
                (
                    row[branch].get("model"),
                    row[branch].get("prompt_tokens") or 0,
                    row[branch].get("completion_tokens") or 0,
                )
            )
    token_groups.extend(
        (
            row.get("model"),
            row.get("prompt_tokens") or 0,
            row.get("completion_tokens") or 0,
        )
        for row in judge.values()
    )
    for model, prompt_tokens, completion_tokens in token_groups:
        rates = price_models.get(model or "", {})
        if rates.get("input_per_million") is not None and rates.get("output_per_million") is not None:
            model_cost = prompt_tokens * float(rates["input_per_million"]) / 1_000_000
            model_cost += completion_tokens * float(rates["output_per_million"]) / 1_000_000
            usage_cost += model_cost
            token_cost_by_model[model or "unknown"] += model_cost
            priced = True
    stt_rates = price_models.get("gpt-4o-mini-transcribe", {})
    eligible_duration_minutes = sum(
        float(row["duration_sec"]) / 60 for row in evaluation_manifest
    )
    if stt_rates.get("per_minute") is not None:
        stt_cost = eligible_duration_minutes * float(stt_rates["per_minute"])
        usage_cost += stt_cost
        priced = True

    system = {
        "stt": {
            "samples": len(raw),
            "audio_minutes": eligible_duration_minutes,
            "failure_rate": sum(row.get("status") != "success" for row in raw.values()) / max(1, len(raw)),
            "p50_latency_ms": _percentile([row.get("latency_ms", 0) for row in raw.values()], 0.5),
            "p95_latency_ms": _percentile([row.get("latency_ms", 0) for row in raw.values()], 0.95),
        },
        "correction": {
            "samples": len(corrected),
            "fallback_rate": sum(row.get("status") == "fallback" for row in corrected.values()) / max(1, len(corrected)),
            "p50_latency_ms": _percentile([row.get("latency_ms", 0) for row in corrected.values()], 0.5),
            "p95_latency_ms": _percentile([row.get("latency_ms", 0) for row in corrected.values()], 0.95),
            "prompt_tokens": sum(row.get("prompt_tokens") or 0 for row in corrected.values()),
            "completion_tokens": sum(row.get("completion_tokens") or 0 for row in corrected.values()),
        },
        "answers": {
            branch: {
                "samples": len(answers),
                "p50_latency_ms": _percentile([row[branch].get("latency_ms", 0) for row in answers.values()], 0.5),
                "p95_latency_ms": _percentile([row[branch].get("latency_ms", 0) for row in answers.values()], 0.95),
                "prompt_tokens": sum(row[branch].get("prompt_tokens") or 0 for row in answers.values()),
                "completion_tokens": sum(row[branch].get("completion_tokens") or 0 for row in answers.values()),
            }
            for branch in ("baseline", "proposed")
        },
        "judge": {
            "samples": len(judge),
            "prompt_tokens": sum(row.get("prompt_tokens") or 0 for row in judge.values()),
            "completion_tokens": sum(
                row.get("completion_tokens") or 0 for row in judge.values()
            ),
        },
    }

    exclusion_counts = Counter(
        row["exclusion_reason"]
        for row in manifest
        if row["eligibility_status"] == "excluded"
    )
    eligible_speakers = Counter(
        row["speaker"]
        for row in manifest
        if row["eligibility_status"] == "eligible"
    )
    summary = {
        "dataset_version": config.get("dataset", {}).get("version"),
        "dataset_audit": {
            "manifest_rows": len(manifest),
            "eligible_samples": sum(
                row["eligibility_status"] == "eligible" for row in manifest
            ),
            "eligible_audio_id_range": [101, 230],
            "eligible_speakers": dict(eligible_speakers),
            "exclusions": dict(exclusion_counts),
            "alignment_validation": {
                "seed": config.get("dataset", {}).get("alignment_validation_seed"),
                "samples": config.get("dataset", {}).get("alignment_validation_samples"),
                "mean_wer": config.get("dataset", {}).get("alignment_validation_mean_wer"),
            },
        },
        "evaluation_split": evaluation_split,
        "expected_samples": len(evaluation_manifest),
        "expected_test_samples": (
            int(config.get("dataset", {}).get("test_samples", 104))
            if evaluation_split == "test"
            else None
        ),
        "observed_transcript_samples": len(transcript_rows),
        "transcript": transcript_summary,
        "retrieval_proxy": {"samples": len(proxy_rows), **proxy_summary},
        "retrieval_gold": {"samples": len(true_rows), **true_summary},
        "human_answers": human_summary,
        "llm_judge_secondary": judge_summary,
        "estimated_token_cost": {
            "currency": pricing.get("currency", "USD"),
            "as_of": pricing.get("as_of"),
            "source": pricing.get("source"),
            "value": usage_cost if priced else None,
            "assumption": pricing.get("assumption"),
            "breakdown": {
                "token_models": dict(token_cost_by_model),
                "gpt-4o-mini-transcribe": stt_cost if priced else None,
            },
        },
        "system": system,
        "limitations": [
            "Benchmark v1 contains 130 aligned samples from two speakers only (Toàn and Trí, 65 each).",
            "Headline test conclusions are based on N=104 and may have wider confidence intervals and lower statistical power than the superseded 182-sample design.",
            "A non-significant p-value must not be interpreted as evidence of no effect; confidence intervals and effect sizes must be considered together.",
            "IDs 1-100 were excluded after detecting audio-reference mismatch and require an independently human-verified mapping before any future reuse.",
            "Proxy retrieval compares against retrieval from the reference transcript and is not true relevance ground truth.",
            "Gold retrieval metrics are emitted only after gold_adjudicated.xlsx is supplied.",
            "Human answer metrics are emitted after both blinded grading workbooks are completed and imported.",
        ],
    }
    atomic_write_json(output_dir / "metrics.json", summary)
    atomic_write_json(output_dir / "transcript_per_sample.json", transcript_rows)
    atomic_write_json(output_dir / "retrieval_proxy_per_sample.json", proxy_rows)
    atomic_write_json(output_dir / "retrieval_gold_per_sample.json", true_rows)
    return summary


def main() -> None:
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("--config", default="dataset_benchmark/benchmark_config.json")
    args = parser.parse_args()
    evaluate(load_config(args.config))


if __name__ == "__main__":
    main()
