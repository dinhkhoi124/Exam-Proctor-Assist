from __future__ import annotations

import csv
import json
from pathlib import Path
import random

try:
    from .common import atomic_write_json, read_jsonl
except ImportError:
    from common import atomic_write_json, read_jsonl


def _openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "openpyxl is required; install requirements-benchmark.txt"
        ) from exc


def read_manifest(path: Path) -> list[dict]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _style_sheet(sheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = {"A": 10, "B": 14, "C": 70, "D": 45, "E": 18, "F": 70}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def balanced_swap_flags(count: int, rng: random.Random) -> list[bool]:
    flags = [False] * (count // 2) + [True] * (count - count // 2)
    rng.shuffle(flags)
    return flags


def _normalize_label(value: object) -> str:
    return " ".join(str(value or "").casefold().split())


def _format_evidence(pages: list[dict], max_chars: int = 32000) -> str:
    blocks = [
        f"[SOURCE: {page.get('source')}, PAGE: {page.get('page')}]\n"
        f"{page.get('content') or ''}"
        for page in pages
    ]
    value = "\n\n".join(blocks)
    return value if len(value) <= max_chars else value[:max_chars] + "\n[TRUNCATED]"


def _add_answer_validations(sheet, row_count: int, headers: list[str]) -> None:
    if row_count <= 0:
        return
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    positions = {header: index + 1 for index, header in enumerate(headers)}
    rubric = DataValidation(type="list", formula1='"1,2,3,4,5"')
    binary = DataValidation(type="list", formula1='"0,1"')
    sheet.add_data_validation(rubric)
    sheet.add_data_validation(binary)
    for label in ("A", "B"):
        for name in ("correctness", "faithfulness", "completeness", "citation"):
            column = get_column_letter(positions[f"{name}_{label}"])
            rubric.add(f"{column}2:{column}{row_count + 1}")
        for name in ("source_page_correct", "task_success"):
            column = get_column_letter(positions[f"{name}_{label}"])
            binary.add(f"{column}2:{column}{row_count + 1}")


def create_gold_templates(
    manifest_path: Path, output_dir: Path, *, overwrite: bool = False
) -> None:
    openpyxl = _openpyxl()
    rows = [row for row in read_manifest(manifest_path) if row["human_eval"] == "True"]
    output_dir.mkdir(parents=True, exist_ok=True)
    for rater in ("A", "B"):
        target = output_dir / f"gold_rater_{rater}.xlsx"
        if target.exists() and not overwrite:
            continue
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Gold"
        sheet.append(
            [
                "audio_id",
                "speaker",
                "reference_transcript",
                "gold_source",
                "gold_pages",
                "expected_key_points",
                "notes",
            ]
        )
        for row in rows:
            sheet.append(
                [
                    int(row["audio_id"]),
                    row["speaker"],
                    row["reference_transcript"],
                    "",
                    "",
                    "",
                    "",
                ]
            )
        _style_sheet(sheet)
        instructions = workbook.create_sheet("Instructions")
        instructions.append(
            [
                "Điền độc lập source PDF, các page phân tách bằng dấu ; và "
                "expected key points. Không tham khảo workbook của người chấm còn lại."
            ]
        )
        workbook.save(target)


def create_answer_templates(
    manifest_path: Path,
    answers_path: Path,
    retrieval_path: Path,
    output_dir: Path,
    seed: int = 43,
    *,
    overwrite: bool = False,
) -> None:
    openpyxl = _openpyxl()
    manifest = {
        int(row["audio_id"]): row
        for row in read_manifest(manifest_path)
        if row["human_eval"] == "True"
    }
    answers = {int(row["audio_id"]): row for row in read_jsonl(answers_path)}
    retrieval = {int(row["audio_id"]): row for row in read_jsonl(retrieval_path)}
    adjudicated_path = output_dir / "gold_adjudicated.xlsx"
    gold = {}
    if adjudicated_path.exists():
        for row in _worksheet_rows(adjudicated_path, "Adjudication"):
            status = _normalize_label(row.get("agreement_status"))
            if row.get("audio_id") is not None and status in {"agreed", "adjudicated"}:
                gold[int(row["audio_id"])] = row
    output_dir.mkdir(parents=True, exist_ok=True)

    for rater_index, rater in enumerate(("A", "B")):
        workbook_target = output_dir / f"answers_rater_{rater}.xlsx"
        mapping_target = output_dir / f"answers_mapping_{rater}.json"
        if workbook_target.exists() and mapping_target.exists() and not overwrite:
            continue
        rng = random.Random(seed + rater_index)
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Blind grading"
        headers = [
            "audio_id",
            "reference_transcript",
            "gold_source",
            "gold_pages",
            "expected_key_points",
            "evidence_A",
            "answer_A",
            "evidence_B",
            "answer_B",
            "correctness_A",
            "correctness_B",
            "faithfulness_A",
            "faithfulness_B",
            "completeness_A",
            "completeness_B",
            "citation_A",
            "citation_B",
            "source_page_correct_A",
            "source_page_correct_B",
            "task_success_A",
            "task_success_B",
            "notes",
        ]
        sheet.append(headers)
        mapping = []
        audio_ids = sorted(set(manifest) & set(answers) & set(retrieval))
        swap_flags = balanced_swap_flags(len(audio_ids), rng)
        for audio_id, swapped in zip(audio_ids, swap_flags):
            answer = answers[audio_id]
            baseline = answer["baseline"]["answer"]
            proposed = answer["proposed"]["answer"]
            baseline_evidence = _format_evidence(
                retrieval[audio_id]["retrieval"]["baseline"]["final_pages"]
            )
            proposed_evidence = _format_evidence(
                retrieval[audio_id]["retrieval"]["proposed"]["final_pages"]
            )
            first, second = (proposed, baseline) if swapped else (baseline, proposed)
            first_evidence, second_evidence = (
                (proposed_evidence, baseline_evidence)
                if swapped
                else (baseline_evidence, proposed_evidence)
            )
            gold_row = gold.get(audio_id, {})
            sheet.append(
                [
                    audio_id,
                    manifest[audio_id]["reference_transcript"],
                    gold_row.get("gold_source", ""),
                    gold_row.get("gold_pages", ""),
                    gold_row.get("expected_key_points", ""),
                    first_evidence,
                    first,
                    second_evidence,
                    second,
                ]
            )
            mapping.append(
                {
                    "audio_id": audio_id,
                    "A": "proposed" if swapped else "baseline",
                    "B": "baseline" if swapped else "proposed",
                }
            )
        sheet.freeze_panes = "A2"
        sheet.column_dimensions["B"].width = 55
        for column in ("C", "D", "E"):
            sheet.column_dimensions[column].width = 45
        for column in ("F", "G", "H", "I"):
            sheet.column_dimensions[column].width = 80
        _add_answer_validations(sheet, len(audio_ids), headers)
        instructions = workbook.create_sheet("Instructions")
        instructions.append(
            [
                "Chấm độc lập 1–5 cho các rubric; task_success dùng 0/1. "
                "Không suy đoán hoặc tìm cách nhận diện pipeline."
            ]
        )
        workbook.save(workbook_target)
        atomic_write_json(mapping_target, mapping)


def _worksheet_rows(path: Path, sheet_name: str) -> list[dict]:
    openpyxl = _openpyxl()
    sheet = openpyxl.load_workbook(path, read_only=True, data_only=True)[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value) if value is not None else "" for value in rows[0]]
    return [dict(zip(headers, values)) for values in rows[1:]]


def create_gold_adjudication(output_dir: Path) -> Path:
    openpyxl = _openpyxl()
    rater_a = {
        int(row["audio_id"]): row
        for row in _worksheet_rows(output_dir / "gold_rater_A.xlsx", "Gold")
        if row.get("audio_id") is not None
    }
    rater_b = {
        int(row["audio_id"]): row
        for row in _worksheet_rows(output_dir / "gold_rater_B.xlsx", "Gold")
        if row.get("audio_id") is not None
    }
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Adjudication"
    headers = [
        "audio_id",
        "reference_transcript",
        "rater_A_source",
        "rater_A_pages",
        "rater_A_key_points",
        "rater_B_source",
        "rater_B_pages",
        "rater_B_key_points",
        "gold_source",
        "gold_pages",
        "expected_key_points",
        "agreement_status",
        "adjudication_notes",
    ]
    sheet.append(headers)
    for audio_id in sorted(set(rater_a) | set(rater_b)):
        first, second = rater_a.get(audio_id, {}), rater_b.get(audio_id, {})
        source_a, source_b = first.get("gold_source"), second.get("gold_source")
        pages_a, pages_b = first.get("gold_pages"), second.get("gold_pages")
        points_a = first.get("expected_key_points")
        points_b = second.get("expected_key_points")
        source_agreed = bool(source_a) and _normalize_label(source_a) == _normalize_label(source_b)
        pages_agreed = bool(pages_a) and _normalize_label(pages_a) == _normalize_label(pages_b)
        points_agreed = bool(points_a) and _normalize_label(points_a) == _normalize_label(points_b)
        agreed = source_agreed and pages_agreed and points_agreed
        sheet.append(
            [
                audio_id,
                first.get("reference_transcript") or second.get("reference_transcript"),
                source_a,
                pages_a,
                points_a,
                source_b,
                pages_b,
                points_b,
                source_a if source_agreed else "",
                pages_a if pages_agreed else "",
                points_a if points_agreed else "",
                "agreed" if agreed else "needs_adjudication",
                "",
            ]
        )
    sheet.freeze_panes = "A2"
    sheet.column_dimensions["B"].width = 65
    sheet.column_dimensions["E"].width = 65
    sheet.column_dimensions["H"].width = 65
    sheet.column_dimensions["K"].width = 65
    from openpyxl.worksheet.datavalidation import DataValidation

    status_validation = DataValidation(
        type="list", formula1='"agreed,adjudicated,unresolved"'
    )
    sheet.add_data_validation(status_validation)
    status_validation.add(f"L2:L{sheet.max_row}")
    target = output_dir / "gold_adjudicated.xlsx"
    workbook.save(target)
    return target


def load_human_grades(output_dir: Path) -> tuple[dict, dict]:
    grades = {}
    mappings = {}
    for rater in ("A", "B"):
        workbook_path = output_dir / f"answers_rater_{rater}.xlsx"
        mapping_path = output_dir / f"answers_mapping_{rater}.json"
        if not workbook_path.exists() or not mapping_path.exists():
            continue
        mapping = {
            int(row["audio_id"]): row
            for row in json.loads(mapping_path.read_text(encoding="utf-8"))
        }
        mappings[rater] = mapping
        for row in _worksheet_rows(workbook_path, "Blind grading"):
            if row.get("audio_id") is None:
                continue
            audio_id = int(row["audio_id"])
            for blind_label in ("A", "B"):
                branch = mapping[audio_id][blind_label]
                grades[(rater, audio_id, branch)] = {
                    "correctness": row.get(f"correctness_{blind_label}"),
                    "faithfulness": row.get(f"faithfulness_{blind_label}"),
                    "completeness": row.get(f"completeness_{blind_label}"),
                    "citation": row.get(f"citation_{blind_label}"),
                    "source_page_correct": row.get(
                        f"source_page_correct_{blind_label}"
                    ),
                    "task_success": row.get(f"task_success_{blind_label}"),
                }
    return grades, mappings


def main() -> None:
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("--annotation-dir", default="dataset_benchmark/annotations")
    parser.add_argument("--build-gold-adjudication", action="store_true")
    args = parser.parse_args()
    if args.build_gold_adjudication:
        print(create_gold_adjudication(Path(args.annotation_dir)))


if __name__ == "__main__":
    main()
