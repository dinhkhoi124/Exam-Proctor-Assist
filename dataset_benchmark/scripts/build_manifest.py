from __future__ import annotations

import argparse
from collections import Counter, defaultdict
import csv
import math
from pathlib import Path
import random
import re
import wave

try:
    from .common import resolve_path
except ImportError:
    from common import resolve_path


TRAILING_ARTIFACT = re.compile(r'\s*"?\}?\s*$')
DATASET_VERSION = "v1-aligned-101-230"
ALIGNED_AUDIO_ID_START = 101
ALIGNED_AUDIO_ID_END = 230
EXPECTED_ELIGIBLE = 130
DEV_COUNT = 26
HUMAN_EVAL_COUNT = 60


def clean_reference(value: object) -> str:
    text = "" if value is None else str(value).strip()
    if text.endswith('"}'):
        text = text[:-2].rstrip()
    elif text.endswith('"'):
        text = text[:-1].rstrip()
    return text


def read_ground_truth(path: Path) -> list[tuple[object, object]]:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        return [(sheet.cell(row, 1).value, sheet.cell(row, 2).value) for row in range(1, 300)]
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "openpyxl is required; install requirements-benchmark.txt"
        ) from exc


def wav_duration(path: Path) -> float:
    with wave.open(str(path), "rb") as audio:
        return audio.getnframes() / audio.getframerate()


def normalized_duplicate_key(text: str) -> str:
    return re.sub(r"\s+", " ", text.casefold()).strip()


def _allocate(groups: dict[str, list[dict]], count: int) -> dict[str, int]:
    total = sum(len(items) for items in groups.values())
    raw = {key: len(items) * count / total for key, items in groups.items()}
    allocation = {key: math.floor(value) for key, value in raw.items()}
    remainder = count - sum(allocation.values())
    order = sorted(groups, key=lambda key: (raw[key] - allocation[key], key), reverse=True)
    for key in order[:remainder]:
        allocation[key] += 1
    return allocation


def stratified_select(rows: list[dict], count: int, seed: int) -> set[int]:
    groups: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        groups[row["stratum"]].append(row)
    allocation = _allocate(groups, count)
    rng = random.Random(seed)
    selected: set[int] = set()
    for key in sorted(groups):
        items = list(groups[key])
        rng.shuffle(items)
        selected.update(int(item["audio_id"]) for item in items[: allocation[key]])
    if len(selected) != count:
        raise RuntimeError(f"Expected {count} selected rows, got {len(selected)}")
    return selected


def build_manifest(excel_path: Path, audio_dir: Path, output_path: Path) -> list[dict]:
    ground_truth = read_ground_truth(excel_path)
    records = []
    for audio_id, (reference_value, speaker_value) in enumerate(ground_truth, start=1):
        audio_path = audio_dir / f"{audio_id}.wav"
        reference = clean_reference(reference_value)
        has_audio = audio_path.exists()
        has_reference = bool(reference)
        if not has_audio:
            status, reason = "excluded", "missing_audio"
        elif not has_reference:
            status, reason = "excluded", "missing_reference"
        elif not ALIGNED_AUDIO_ID_START <= audio_id <= ALIGNED_AUDIO_ID_END:
            status, reason = "excluded", "audio_reference_mismatch"
        else:
            status, reason = "eligible", ""
        records.append(
            {
                "audio_id": audio_id,
                "audio_path": str(audio_path.relative_to(resolve_path("."))).replace("\\", "/") if has_audio else "",
                "speaker": "" if speaker_value is None else str(speaker_value).strip(),
                "duration_sec": round(wav_duration(audio_path), 3) if has_audio else "",
                "reference_transcript": reference,
                "eligibility_status": status,
                "exclusion_reason": reason,
                "duplicate_group": "",
                "split": "excluded",
                "human_eval": False,
                "gold_source": "",
                "gold_pages": "",
                "expected_key_points": "",
            }
        )

    duplicate_counts = Counter(
        normalized_duplicate_key(row["reference_transcript"])
        for row in records
        if row["reference_transcript"]
    )
    duplicate_ids = {
        key: f"dup-{index:02d}"
        for index, key in enumerate(
            sorted(key for key, count in duplicate_counts.items() if count > 1), start=1
        )
    }
    for row in records:
        key = normalized_duplicate_key(row["reference_transcript"])
        row["duplicate_group"] = duplicate_ids.get(key, "")

    eligible = [row for row in records if row["eligibility_status"] == "eligible"]
    duration_order = sorted(eligible, key=lambda row: (float(row["duration_sec"]), row["audio_id"]))
    quartiles = {}
    for index, row in enumerate(duration_order):
        quartiles[int(row["audio_id"])] = min(3, index * 4 // len(duration_order)) + 1
    for row in eligible:
        row["stratum"] = f"{row['speaker']}|q{quartiles[int(row['audio_id'])]}"

    if len(eligible) != EXPECTED_ELIGIBLE:
        raise RuntimeError(
            f"{DATASET_VERSION} expected {EXPECTED_ELIGIBLE} eligible rows, "
            f"got {len(eligible)}"
        )
    dev_ids = stratified_select(eligible, count=DEV_COUNT, seed=42)
    for row in eligible:
        row["split"] = "dev" if int(row["audio_id"]) in dev_ids else "test"
    test_rows = [row for row in eligible if row["split"] == "test"]
    human_ids = stratified_select(test_rows, count=HUMAN_EVAL_COUNT, seed=43)
    for row in eligible:
        row["human_eval"] = int(row["audio_id"]) in human_ids
        row.pop("stratum", None)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = list(records[0].keys())
    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(records)
    return records


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", default="dataset_benchmark/AUDIO_299.xlsx")
    parser.add_argument(
        "--audio-dir",
        default="dataset_benchmark/Audio_wav-20260708T055213Z-3-001/Audio_wav",
    )
    parser.add_argument("--output", default="dataset_benchmark/manifest.csv")
    args = parser.parse_args()
    records = build_manifest(
        resolve_path(args.excel), resolve_path(args.audio_dir), resolve_path(args.output)
    )
    eligible = [row for row in records if row["eligibility_status"] == "eligible"]
    print(
        f"Wrote {len(records)} rows: {len(eligible)} eligible, "
        f"{sum(row['split'] == 'dev' for row in eligible)} dev, "
        f"{sum(row['split'] == 'test' for row in eligible)} test"
    )


if __name__ == "__main__":
    main()
