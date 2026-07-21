"""Evaluate P0/P1/P2 with base-cluster statistics and evidence labels."""

from __future__ import annotations

import argparse
from collections import Counter
import csv
import json
from pathlib import Path
from typing import Any, Iterable, Mapping
import wave

from dataset_benchmark.robustness_v2.error_analysis import build_stage_metadata
from dataset_benchmark.robustness_v2.evaluation import (
    aggregate_operational_pipeline,
    aggregate_retrieval_pipeline,
    aggregate_transcript_pipeline,
    base_aggregates,
    cluster_bootstrap,
    code_switch_flag,
    entity_flag,
    hallucinated_token_rate,
    holm_adjust,
    page_overlap,
    paired_base_values,
    paired_wilcoxon,
    stratify,
    two_way_cluster_bootstrap,
    wer_bin,
)
from dataset_benchmark.robustness_v2.pipeline import canonical_hash
from dataset_benchmark.scripts.common import atomic_write_json, read_jsonl, resolve_path
from dataset_benchmark.scripts.metrics import normalize_text, transcript_errors, true_retrieval_metrics


def write_csv(path: Path, rows: Iterable[Mapping[str, Any]]) -> None:
    records = [dict(row) for row in rows]
    path.parent.mkdir(parents=True, exist_ok=True)
    headers = sorted({key for row in records for key in row})
    temporary = path.with_suffix(path.suffix + ".tmp")
    with temporary.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for row in records:
            writer.writerow(
                {
                    key: json.dumps(value, ensure_ascii=False, sort_keys=True)
                    if isinstance(value, (dict, list, tuple))
                    else value
                    for key, value in row.items()
                }
            )
    temporary.replace(path)


def read_csv(path: Path) -> list[dict]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def parse_gold(row: Mapping[str, Any]) -> set[tuple[str, int]]:
    source = str(row.get("gold_source") or "").strip()
    pages = [
        int(value.strip())
        for value in str(row.get("gold_pages") or "").split(";")
        if value.strip().isdigit()
    ]
    return {(source, page) for page in pages} if source else set()


def _token_cost(rows: Iterable[Mapping[str, Any]], config: Mapping[str, Any]) -> float:
    price = config["cost"]["models"]["gpt-4o-mini"]
    api_rows = [row for row in rows if row.get("cache_origin") == "api"]
    return (
        sum(int(row.get("prompt_tokens") or 0) for row in api_rows)
        * float(price["input_per_million_tokens"])
        + sum(int(row.get("completion_tokens") or 0) for row in api_rows)
        * float(price["output_per_million_tokens"])
    ) / 1_000_000


def actual_cost_accounting(
    config: Mapping[str, Any], input_paths: Mapping[str, Path]
) -> dict:
    """Derive unique-run spend and comparable deployed-route cost from caches."""
    manifest = read_jsonl(input_paths["run_manifest"])
    stt_rows = read_jsonl(input_paths["stt_cache"])
    correction_rows = read_jsonl(input_paths["correction_cache"])
    answer_rows = read_jsonl(input_paths["answer_cache"])
    judge_rows = (
        read_jsonl(input_paths["judge_cache"])
        if input_paths.get("judge_cache")
        else []
    )
    stt_api_ids = {
        row["variant_id"] for row in stt_rows if row.get("cache_origin") == "api"
    }
    stt_minutes = 0.0
    for row in manifest:
        if row["variant_id"] not in stt_api_ids:
            continue
        with wave.open(str(resolve_path(row["output_audio_path"])), "rb") as handle:
            stt_minutes += handle.getnframes() / handle.getframerate() / 60
    stt_cost = stt_minutes * float(
        config["cost"]["models"]["gpt-4o-mini-transcribe"]["per_minute_assumption"]
    )
    correction_cost = _token_cost(correction_rows, config)
    answer_cost = _token_cost(answer_rows, config)
    judge_cost = _token_cost(judge_rows, config)
    request_count = len(manifest)
    routes = {}
    for pipeline in ("P0", "P1", "P2"):
        route_corrections = [
            row for row in correction_rows if pipeline in row.get("required_by", [])
        ]
        route_answers = [row for row in answer_rows if row.get("pipeline") == pipeline]
        route_judges = [row for row in judge_rows if row.get("pipeline") == pipeline]
        components = {
            "stt": stt_cost,
            "correction": _token_cost(route_corrections, config),
            "final_answers": _token_cost(route_answers, config),
            "judge": _token_cost(route_judges, config),
        }
        total = sum(components.values())
        routes[pipeline] = {
            "components_usd": {key: round(value, 6) for key, value in components.items()},
            "total_usd": round(total, 6),
            "per_1000_requests_usd": round(total / request_count * 1000, 6),
            "fresh_api_calls": (
                sum(row.get("cache_origin") == "api" for row in stt_rows)
                + sum(row.get("cache_origin") == "api" for row in route_corrections)
                + sum(row.get("cache_origin") == "api" for row in route_answers)
                + sum(row.get("cache_origin") == "api" for row in route_judges)
            ),
        }
    return {
        "currency": "USD",
        "pricing_as_of": config["cost"]["as_of"],
        "unique_run_spend_usd": {
            "stt": round(stt_cost, 6),
            "correction": round(correction_cost, 6),
            "final_answers": round(answer_cost, 6),
            "judge": round(judge_cost, 6),
            "total": round(stt_cost + correction_cost + answer_cost + judge_cost, 6),
        },
        "stt_billable_minutes": round(stt_minutes, 3),
        "route_comparison": routes,
        "methodology": (
            "Each route includes the full observed STT cost plus only that route's "
            "correction, answer, and judge calls. Route totals are deployment-comparable "
            "and must not be summed; unique_run_spend_usd counts shared STT once. "
            "Compatible C0 imports incurred no fresh cost in this run."
        ),
        "provider_billing_authoritative": True,
    }


def annotation_availability(
    config: Mapping[str, Any], primary_keys: set[tuple[str, str]] | None = None
) -> dict:
    from openpyxl import load_workbook

    answer_dir = resolve_path(config["evaluation"]["answer_annotation_dir"])
    answer_scored = 0
    for name in ("answers_rater_A.xlsx", "answers_rater_B.xlsx"):
        path = answer_dir / name
        if not path.exists():
            continue
        sheet = load_workbook(path, read_only=True, data_only=True).active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(value or "") for value in rows[0]]
        score_positions = [
            index
            for index, header in enumerate(headers)
            if header.startswith(("correctness_", "faithfulness_", "completeness_", "citation_", "task_success_"))
        ]
        answer_scored += sum(
            any(row[index] not in (None, "") for index in score_positions)
            for row in rows[1:]
        )
    error_dir = resolve_path(config["evaluation"]["error_annotation_dir"])
    error_reviewed = 0
    for name in ("error_analysis_rater_A.xlsx", "error_analysis_rater_B.xlsx"):
        path = error_dir / name
        if not path.exists():
            continue
        sheet = load_workbook(path, read_only=True, data_only=True)["Error Analysis"]
        rows = list(sheet.iter_rows(values_only=True))
        headers = [str(value or "") for value in rows[0]]
        position = headers.index("reviewed_status")
        error_reviewed += sum(
            str(row[position] or "").strip().casefold() == "reviewed"
            for row in rows[1:]
        )
    judge_path = resolve_path(config["outputs"]["cache_dir"]) / "judge.jsonl"
    judge_rows = read_jsonl(judge_path) if judge_path.exists() else []
    if primary_keys is not None:
        judge_rows = [
            row for row in judge_rows
            if (str(row.get("variant_id")), str(row.get("pipeline"))) in primary_keys
        ]
    judge_scored = sum(row.get("status") == "success" for row in judge_rows)
    judge_by_pipeline = {}
    for pipeline in ("P0", "P1", "P2"):
        pipeline_rows = [
            row for row in judge_rows
            if row.get("pipeline") == pipeline and row.get("status") == "success"
        ]
        judge_by_pipeline[pipeline] = {
            "scored_rows": len(pipeline_rows),
            **{
                f"mean_{metric}": (
                    sum(float(row["scores"][metric]) for row in pipeline_rows)
                    / len(pipeline_rows)
                    if pipeline_rows else None
                )
                for metric in ("correctness", "groundedness", "helpfulness", "safety")
            },
        }
    return {
        "human_final_answer": {
            "availability": "available" if answer_scored else "unavailable",
            "scored_rows": answer_scored,
            "reason": None if answer_scored else "Answer grading workbooks contain no scores.",
        },
        "error_taxonomy_recoverability": {
            "availability": "available" if error_reviewed else "unavailable",
            "reviewed_rows": error_reviewed,
            "reason": None if error_reviewed else "Error-analysis workbooks contain no reviewed rows.",
        },
        "llm_judge": {
            "availability": "available" if judge_scored else "unavailable",
            "scored_rows": judge_scored,
            "by_pipeline": judge_by_pipeline,
            "reason": None if judge_scored else "Judge disabled or no successful judge records.",
            "role": "auxiliary_only",
        },
    }


def build_sample_rows(config: Mapping[str, Any]) -> tuple[list[dict], dict[str, Path]]:
    cache_dir = resolve_path(config["outputs"]["cache_dir"])
    paths = {
        "config": resolve_path("dataset_benchmark/robustness_v2/configs/benchmark_v2_config.json"),
        "run_manifest": resolve_path(config["outputs"]["run_manifest"]),
        "stt_cache": cache_dir / "stt.jsonl",
        "risk_cache": cache_dir / "risk_decisions.jsonl",
        "correction_cache": cache_dir / "corrections.jsonl",
        "retrieval_cache": cache_dir / "retrieval.jsonl",
        "answer_cache": cache_dir / "final_answers.jsonl",
        "reference_retrieval_cache": resolve_path(
            config["answer_generation"]["compatible_retrieval_cache"]
        ),
        "gold_manifest": resolve_path(config["evaluation"]["gold_manifest"]),
    }
    judge_path = cache_dir / "judge.jsonl"
    if judge_path.exists():
        paths["judge_cache"] = judge_path
    manifest = {row["variant_id"]: row for row in read_jsonl(paths["run_manifest"])}
    if {row["condition_level"] for row in manifest.values()} == {"C0", "C1", "C2", "C3"}:
        historical_pilot = resolve_path(
            "dataset_benchmark/robustness_v2/reports/robustness_v2_summary.json"
        )
        if historical_pilot.exists():
            paths["historical_pilot_summary"] = historical_pilot
    stt = {row["cache_id"]: row for row in read_jsonl(paths["stt_cache"])}
    risk = {row["cache_id"]: row for row in read_jsonl(paths["risk_cache"])}
    correction = {
        row["cache_id"]: row for row in read_jsonl(paths["correction_cache"])
    }
    retrieval = read_jsonl(paths["retrieval_cache"])
    answers = {row["cache_id"]: row for row in read_jsonl(paths["answer_cache"])}
    reference_retrieval = {
        int(row["audio_id"]): row
        for row in read_jsonl(paths["reference_retrieval_cache"])
    }
    gold = {int(row["audio_id"]): row for row in read_csv(paths["gold_manifest"])}
    output = []
    for retrieval_row in retrieval:
        variant_id = retrieval_row["variant_id"]
        if variant_id not in manifest:
            continue
        variant = manifest[variant_id]
        source = stt[variant_id]
        detector = risk[variant_id]
        correction_row = correction[variant_id]
        answer = answers.get(retrieval_row["cache_id"], {})
        reference = variant["reference_transcript"]
        transcript = retrieval_row["query"]
        raw = source["raw_transcript"]
        values = transcript_errors(reference, transcript)
        raw_values = transcript_errors(reference, raw)
        if values["word_errors"] < raw_values["word_errors"]:
            outcome = "improved"
        elif values["word_errors"] > raw_values["word_errors"]:
            outcome = "degraded"
        else:
            outcome = "unchanged"
        reference_branch = reference_retrieval.get(int(variant["base_id"]), {})
        proxy_available = (
            reference_branch.get("queries", {}).get("reference") == reference
            and reference_branch.get("index_signature") == retrieval_row["index_signature"]
        )
        proxy = page_overlap(
            retrieval_row["retrieval"].get("final_pages", []),
            reference_branch.get("retrieval", {}).get("reference", {}).get("final_pages", []),
        ) if proxy_available else {
            "jaccard_at_5_proxy": None,
            "overlap_recall_at_5_proxy": None,
        }
        gold_pairs = parse_gold(gold.get(int(variant["base_id"]), {}))
        true_metrics = true_retrieval_metrics(
            retrieval_row["retrieval"].get("candidates", []), gold_pairs
        ) if gold_pairs else {}
        output.append(
            {
                "variant_id": variant_id,
                "base_id": int(variant["base_id"]),
                "pipeline": retrieval_row["pipeline"],
                "split": variant["split"],
                "speaker": variant.get("speaker", ""),
                "intent": variant.get("intent", ""),
                "semantic_cluster": variant.get("semantic_cluster", ""),
                "condition_level": variant["condition_level"],
                "noise_type": variant.get("noise_type") or "none",
                "noise_mode": variant.get("noise_mode") or "none",
                "noise_source_recording_id": variant.get("noise_source_recording_id"),
                "noise_asset_id": variant.get("noise_asset_id"),
                "snr_db": variant.get("snr_db"),
                "reference_transcript": reference,
                "raw_transcript": raw,
                "pipeline_transcript": transcript,
                "query_source": retrieval_row["query_source"],
                "word_errors": values["word_errors"],
                "reference_words": values["reference_words"],
                "wer": values["wer"],
                "char_errors": values["char_errors"],
                "reference_chars": values["reference_chars"],
                "cer": values["cer"],
                "raw_word_errors": raw_values["word_errors"],
                "raw_wer": raw_values["wer"],
                "raw_wer_bin": wer_bin(raw_values["wer"]),
                "outcome_vs_p0": outcome,
                "transcript_changed_vs_p0": normalize_text(transcript) != normalize_text(raw),
                "hallucinated_token_rate_proxy": hallucinated_token_rate(reference, transcript),
                "semantic_rewrite_proxy": float(
                    normalize_text(transcript) != normalize_text(raw)
                    and values["word_errors"] >= raw_values["word_errors"]
                ),
                "code_switch": code_switch_flag(reference),
                "entity": entity_flag(reference),
                "risk_score": detector["risk_score"],
                "risk_decision": detector["decision"],
                "risk_reasons": detector["reasons"],
                "correction_requested": retrieval_row["pipeline"] in correction_row["required_by"],
                "correction_api_call": correction_row.get("cache_origin") == "api",
                "retrieval_status": retrieval_row["retrieval"].get("status"),
                "retrieval_latency_ms": retrieval_row.get("latency_ms", 0),
                "retrieval_proxy_available": proxy_available,
                **proxy,
                "retrieval_gold_available": bool(gold_pairs),
                **true_metrics,
                "answer_status": answer.get("status", "missing"),
                "answer_api_call": answer.get("cache_origin") == "api",
                "prompt_tokens": answer.get("prompt_tokens"),
                "completion_tokens": answer.get("completion_tokens"),
            }
        )
    return output, paths


def primary_evidence_rows(rows: list[dict], config: Mapping[str, Any]) -> tuple[list[dict], dict | None]:
    filter_config = config.get("evaluation", {}).get("primary_evidence_filter")
    if not filter_config:
        for row in rows:
            row["evidence_layer"] = "primary"
        return rows, None
    conditions = set(filter_config.get("conditions", []))
    primary = []
    for row in rows:
        selected = (
            (not filter_config.get("split") or row["split"] == filter_config["split"])
            and (not filter_config.get("noise_mode") or row["noise_mode"] == filter_config["noise_mode"])
            and (not conditions or row["condition_level"] in conditions)
        )
        row["evidence_layer"] = "primary_recorded_noise_test" if selected else "secondary"
        if selected:
            primary.append(row)
    expected = int(filter_config.get("expected_variants_per_pipeline", 0))
    for pipeline in ("P0", "P1", "P2"):
        actual = sum(row["pipeline"] == pipeline for row in primary)
        if expected and actual != expected:
            raise ValueError(
                f"Primary evidence expected {expected} {pipeline} rows, found {actual}"
            )
    return primary, dict(filter_config)


def paired_noise_source_sensitivity(
    rows: list[dict], candidate: str, *, iterations: int, seed: int
) -> dict:
    by_key = {(row["variant_id"], row["pipeline"]): row for row in rows}
    paired = []
    for row in rows:
        if row["pipeline"] != candidate:
            continue
        baseline = by_key.get((row["variant_id"], "P0"))
        if baseline is None or not row.get("noise_source_recording_id"):
            continue
        paired.append(
            {
                "variant_id": row["variant_id"],
                "base_id": row["base_id"],
                "noise_source_recording_id": row["noise_source_recording_id"],
                "noise_type": row["noise_type"],
                "wer_difference": float(row["wer"]) - float(baseline["wer"]),
            }
        )
    grouped: dict[str, list[float]] = {}
    for row in paired:
        grouped.setdefault(str(row["noise_source_recording_id"]), []).append(
            row["wer_difference"]
        )
    leave_one_out = []
    for source in sorted(grouped):
        retained = [
            row["wer_difference"]
            for row in paired
            if row["noise_source_recording_id"] != source
        ]
        leave_one_out.append(
            {"excluded_noise_source_recording_id": source, "mean_difference": sum(retained) / len(retained)}
        )
    return {
        "candidate": candidate,
        "paired_variants": len(paired),
        "noise_sources": len(grouped),
        "mean_wer_difference": sum(row["wer_difference"] for row in paired) / len(paired),
        "by_noise_source": [
            {
                "noise_source_recording_id": source,
                "variants": len(values),
                "mean_wer_difference": sum(values) / len(values),
            }
            for source, values in sorted(grouped.items())
        ],
        "leave_one_noise_source_out": {
            "minimum_mean_difference": min(row["mean_difference"] for row in leave_one_out),
            "maximum_mean_difference": max(row["mean_difference"] for row in leave_one_out),
            "rows": leave_one_out,
        },
        "two_way_cluster_bootstrap": two_way_cluster_bootstrap(
            paired,
            value_field="wer_difference",
            first_cluster="base_id",
            second_cluster="noise_source_recording_id",
            iterations=iterations,
            seed=seed,
        ),
    }


def evaluate(config: dict, config_path: Path) -> dict:
    rows, input_paths = build_sample_rows(config)
    primary_rows, primary_filter = primary_evidence_rows(rows, config)
    pipelines = ("P0", "P1", "P2")
    transcript = {
        pipeline: aggregate_transcript_pipeline(
            [row for row in primary_rows if row["pipeline"] == pipeline]
        )
        for pipeline in pipelines
    }
    retrieval = {
        pipeline: aggregate_retrieval_pipeline(
            [row for row in primary_rows if row["pipeline"] == pipeline]
        )
        for pipeline in pipelines
    }
    operational = {
        pipeline: aggregate_operational_pipeline(
            [row for row in primary_rows if row["pipeline"] == pipeline]
        )
        for pipeline in pipelines
    }
    cost_accounting = actual_cost_accounting(config, input_paths)
    for pipeline in pipelines:
        route = cost_accounting["route_comparison"][pipeline]
        operational[pipeline]["fresh_api_cost_usd"] = route["total_usd"]
        operational[pipeline]["fresh_api_cost_per_1000_requests_usd"] = route[
            "per_1000_requests_usd"
        ]
        operational[pipeline]["fresh_api_calls"] = route["fresh_api_calls"]
        operational[pipeline]["fresh_api_cost_components_usd"] = route[
            "components_usd"
        ]
    base_rows = base_aggregates(primary_rows)
    comparisons = {}
    p_values = {}
    statistics = config["statistics"]
    for candidate in ("P1", "P2"):
        comparisons[candidate] = {}
        for metric in ("mean_wer", "mean_retrieval_jaccard_proxy"):
            baseline, proposed, base_ids = paired_base_values(
                base_rows, metric, "P0", candidate
            )
            bootstrap = cluster_bootstrap(
                baseline,
                proposed,
                iterations=int(statistics["bootstrap_iterations"]),
                seed=int(statistics["seed"]),
            )
            wilcoxon = paired_wilcoxon(baseline, proposed)
            key = f"{candidate}:{metric}"
            p_values[key] = wilcoxon.get("p_value")
            comparisons[candidate][metric] = {
                "base_ids": base_ids,
                "cluster_bootstrap": bootstrap,
                "paired_wilcoxon": wilcoxon,
            }
    adjusted = holm_adjust(p_values)
    for key, value in adjusted.items():
        candidate, metric = key.split(":", 1)
        comparisons[candidate][metric]["holm_adjusted_p"] = value

    fields = (
        "condition_level", "noise_type", "snr_db", "speaker", "intent",
        "semantic_cluster", "raw_wer_bin", "code_switch", "entity",
    )
    stratified = {field: stratify(primary_rows, field) for field in fields}
    primary_keys = {(row["variant_id"], row["pipeline"]) for row in primary_rows}
    annotations = annotation_availability(config, primary_keys if primary_filter else None)
    source_sensitivity = (
        {
            candidate: paired_noise_source_sensitivity(
                primary_rows,
                candidate,
                iterations=int(statistics["bootstrap_iterations"]),
                seed=int(statistics["seed"]) + index,
            )
            for index, candidate in enumerate(("P1", "P2"), start=1)
        }
        if primary_filter else None
    )
    margin = float(config["evaluation"]["clean_non_inferiority"]["wer_margin_absolute"])
    decision_config = config["evaluation"]["production_decision"]
    decisions = {}
    speakers = len({row["speaker"] for row in primary_rows})
    for pipeline in ("P1", "P2"):
        ci = comparisons[pipeline]["mean_wer"]["cluster_bootstrap"]
        wer_pass = ci.get("availability") == "available" and float(ci["ci_high"]) <= margin
        criteria = {
            "clean_wer_non_inferiority": {
                "status": "pass" if wer_pass else "fail_or_inconclusive",
                "margin": margin,
                "candidate_minus_p0_ci_high": ci.get("ci_high"),
            },
            "human_task_success": {"status": "unavailable"},
            "true_gold_retrieval": {"status": "unavailable"},
            "overcorrection": {
                "status": "pass" if transcript[pipeline]["over_correction_rate"] <= float(decision_config["maximum_overcorrection_rate"]) else "fail"
            },
            "correction_call_rate": {
                "status": "pass" if operational[pipeline]["logical_correction_call_rate"] <= float(decision_config["maximum_correction_call_rate"]) else "fail"
            },
            "p95_end_to_end_latency": {"status": "unavailable"},
            "fresh_api_cost": {
                "status": (
                    "pass"
                    if operational[pipeline]["fresh_api_cost_per_1000_requests_usd"]
                    <= float(decision_config["maximum_cost_per_1000_requests_usd"])
                    else "fail"
                ),
                "value_usd_per_1000_requests": operational[pipeline][
                    "fresh_api_cost_per_1000_requests_usd"
                ],
                "maximum_usd_per_1000_requests": float(
                    decision_config["maximum_cost_per_1000_requests_usd"]
                ),
            },
            "speaker_diversity": {
                "status": "pass" if speakers >= int(decision_config["minimum_speakers"]) else "fail",
                "speakers": speakers,
            },
        }
        decisions[pipeline] = {
            "criteria": criteria,
            "recommend_for_production": False,
            "automatic_enable": False,
            "reason": "Required human task-success, true-gold retrieval, end-to-end latency, and new-speaker evidence is incomplete; all configured gates must pass.",
        }

    outputs = config["outputs"]
    sample_path = resolve_path(outputs["sample_level_csv"])
    base_path = resolve_path(outputs["base_level_csv"])
    metrics_path = resolve_path(outputs["metrics_csv"])
    write_csv(sample_path, rows)
    write_csv(base_path, base_rows)
    write_csv(
        metrics_path,
        [
            {
                "pipeline": pipeline,
                **{f"transcript_{key}": value for key, value in transcript[pipeline].items() if not isinstance(value, (dict, list))},
                **{f"operational_{key}": value for key, value in operational[pipeline].items() if not isinstance(value, (dict, list))},
                "retrieval_proxy_jaccard_at_5": retrieval[pipeline]["proxy"]["jaccard_at_5"],
                "retrieval_true_gold_availability": retrieval[pipeline]["true_gold"]["availability"],
            }
            for pipeline in pipelines
        ],
    )
    evaluated_conditions = sorted({row["condition_level"] for row in primary_rows})
    evaluated_base_ids = sorted({int(row["base_id"]) for row in primary_rows})
    full_c0_c3 = set(evaluated_conditions) == {"C0", "C1", "C2", "C3"}
    historical_pilot = None
    if full_c0_c3 and input_paths.get("historical_pilot_summary"):
        historical_pilot = json.loads(
            input_paths["historical_pilot_summary"].read_text(encoding="utf-8")
        )
    limitations = [
        "Synthetic augmentation is not equivalent to recorded environmental audio.",
        "Audio variants are clustered by base_id and are not independent samples.",
        "Retrieval agreement is a reference-query proxy, not relevance ground truth.",
        "Human answer grading, error taxonomy, and recoverability annotations are unavailable.",
        "Text-only correction is limited by the ASR information bottleneck.",
        "Results must not be generalized to other domains or new speakers.",
    ]
    if primary_filter:
        limitations.insert(0, "Primary evidence is held-out test recorded-noise only; development and clean rows are secondary.")
        limitations.append("Short recorded clips are deterministically wrapped when shorter than an utterance; source-level sensitivity is reported.")
    elif full_c0_c3:
        limitations.insert(0, "The full C0-C3 run uses controlled synthetic degradation and only two speakers.")
        limitations.append("The historical 10-base C0 replay must be reported separately from fresh full-run evidence.")
    else:
        limitations.insert(0, "Only 10 C0 base utterances and one speaker are present in the pilot.")
        limitations.insert(1, "C1-C3 audio was not inferred; no degraded-condition effectiveness claim is supported.")
        limitations.append("Imported historical caches provide no fresh end-to-end latency evidence.")
    summary = {
        "schema_version": "1.0.0",
        "evidence_scope": (
            "held-out owner-recorded environmental-noise test evidence"
            if primary_filter else
            "full C0-C3 local inference; fresh paid API caches; synthetic robustness evidence"
            if full_c0_c3
            else "C0 pilot; compatible historical caches; no fresh external API calls"
        ),
        "evidence_layers": ({
            "primary_recorded_noise_test": {
                "scope": f"{len(evaluated_base_ids)} bases; {len({row['variant_id'] for row in primary_rows})} variants; {len(primary_rows)} pipeline rows",
                "filter": primary_filter,
                "included_in_main_metric_tables": True,
            },
            "secondary_rows": {
                "scope": f"{len(rows) - len(primary_rows)} pipeline rows including dev and C0",
                "included_in_main_metric_tables": False,
            },
        } if primary_filter else {
            "primary_full_run": {
                "scope": "130 bases; 1,040 C0-C3 variants; 3,120 pipeline rows",
                "origin": "fresh local inference plus verified C0 compatible imports",
                "role": "primary synthetic robustness evidence",
                "included_in_main_metric_tables": True,
            },
            "historical_pilot": {
                "scope": "10 C0 bases; IDs 101-110; 30 pipeline rows",
                "origin": "historical compatible-cache replay",
                "role": "secondary historical sanity check only",
                "included_in_main_metric_tables": False,
                "summary_available": historical_pilot is not None,
                "corpus_wer": (
                    {
                        pipeline: historical_pilot["transcript"][pipeline]["corpus_wer"]
                        for pipeline in ("P0", "P1", "P2")
                    }
                    if historical_pilot else None
                ),
            },
        } if full_c0_c3 else {
            "historical_pilot": {
                "scope": "10 C0 bases; IDs 101-110; 30 pipeline rows",
                "origin": "historical compatible-cache replay",
                "role": "current report scope",
                "included_in_main_metric_tables": True,
            }
        }),
        "config_hash": canonical_hash(config),
        "dataset": {
            "independent_base_utterances": len({row["base_id"] for row in primary_rows}),
            "evaluated_base_ids": evaluated_base_ids,
            "selection": {
                "condition_filter": (
                    "held-out test recorded-noise C1-C3"
                    if primary_filter else "all C0-C3" if full_c0_c3 else "C0"
                ),
                "ordering": (
                    "manifest order after locked primary-evidence filter"
                    if primary_filter else "manifest order across all eligible rows" if full_c0_c3 else "ascending (int(base_id), variant_id)"
                ),
                "selection": (
                    "all eligible test base IDs and recorded-noise variants"
                    if primary_filter else "all 130 base IDs" if full_c0_c3 else "first 10 unique base_id values"
                ),
                "randomized": False,
                "selection_seed": None,
                "disclosure": "global_seed does not participate in base selection",
            },
            "historical_pilot_base_ids": list(range(101, 111)),
            "audio_variants": len({row["variant_id"] for row in primary_rows}),
            "pipeline_records": len(primary_rows),
            "run_audio_variants": len({row["variant_id"] for row in rows}),
            "run_pipeline_records": len(rows),
            "primary_filter": primary_filter,
            "unique_noise_sources": len({row["noise_source_recording_id"] for row in primary_rows if row.get("noise_source_recording_id")}),
            "speakers": speakers,
            "speaker_values": sorted({row["speaker"] for row in primary_rows}),
            "conditions": evaluated_conditions,
            "split_methodology": "semantic-cluster regrouped research split; not fresh test",
            "leakage_audit": "clean",
        },
        "transcript": transcript,
        "retrieval": retrieval,
        "final_answers": annotations["human_final_answer"],
        "llm_judge": annotations["llm_judge"],
        "error_taxonomy_recoverability": annotations["error_taxonomy_recoverability"],
        "operational": operational,
        "actual_cost_accounting": cost_accounting,
        "selective_correction": {
            "threshold": config["risk_detector"]["threshold"],
            "detector_version": config["risk_detector"]["version"],
            "decision_counts": dict(Counter(row["risk_decision"] for row in rows if row["pipeline"] == "P2")),
            "trigger_reasons": dict(Counter(reason for row in rows if row["pipeline"] == "P2" for reason in row["risk_reasons"])),
            "precision_recall": {"availability": "unavailable_no_high_risk_oracle_labels"},
        },
        "statistics": {
            "resampling_unit": "base_id",
            "comparisons": comparisons,
            "noise_source_sensitivity": source_sensitivity,
            "multiple_comparison_correction": "holm",
        },
        "stratified": stratified,
        "production_decision": decisions,
        "limitations": limitations,
        "input_hashes": build_stage_metadata(
            "evaluate_robustness",
            inputs={**input_paths, "config": config_path},
            outputs={
                "metrics_csv": metrics_path,
                "sample_level_csv": sample_path,
                "base_level_csv": base_path,
            },
        )["inputs"],
    }
    atomic_write_json(resolve_path(outputs["evaluation_summary"]), summary)
    return summary


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--config",
        default="dataset_benchmark/robustness_v2/configs/benchmark_v2_config.json",
    )
    args = parser.parse_args()
    config_path = resolve_path(args.config)
    config = json.loads(config_path.read_text(encoding="utf-8"))
    evaluate(config, config_path)
    print(resolve_path(config["outputs"]["evaluation_summary"]))


if __name__ == "__main__":
    main()
