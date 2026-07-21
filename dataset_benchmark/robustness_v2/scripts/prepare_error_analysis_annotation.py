"""Create blinded error-analysis workbooks and an adjudication workbook."""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
import random
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from dataset_benchmark.robustness_v2.error_analysis import (
    ERROR_TAXONOMY,
    RECOVERABILITY_LEVELS,
    build_stage_metadata,
    build_annotation_records,
    wer_bin,
)
from dataset_benchmark.scripts.common import atomic_write_json, read_jsonl, resolve_path


HEADERS = [
    "sample_id",
    "speaker",
    "intent",
    "reference",
    "raw_transcript",
    "corrected_transcript",
    "error_id",
    "error_source",
    "error_type",
    "reference_span",
    "hypothesis_span",
    "reference_start",
    "reference_end",
    "hypothesis_start",
    "hypothesis_end",
    "edit_count",
    "wer_edit_count",
    "correction_resolved",
    "reference_word_count",
    "raw_word_errors",
    "corrected_word_errors",
    "raw_wer",
    "corrected_wer",
    "raw_wer_bin",
    "transcript_changed",
    "primary_taxonomy",
    "secondary_tags",
    "text_recoverability",
    "notes",
    "annotator",
    "reviewed_status",
]


def _read_manifest(path: Path) -> list[dict]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _config(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def _annotation_values(record: dict) -> list:
    values = dict(record)
    values["error_type"] = values.pop("type")
    values["raw_wer_bin"] = wer_bin(float(values["raw_wer"]))
    values.update(
        {
            "primary_taxonomy": "",
            "secondary_tags": "",
            "text_recoverability": "",
            "notes": "",
            "annotator": "",
            "reviewed_status": "",
        }
    )
    return [values.get(header, "") for header in HEADERS]


def _style_annotation_sheet(sheet, row_count: int) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    wide = {
        "D": 55,
        "E": 55,
        "F": 55,
        "J": 25,
        "K": 25,
        "AA": 30,
        "AC": 45,
    }
    for index in range(1, len(HEADERS) + 1):
        letter = get_column_letter(index)
        sheet.column_dimensions[letter].width = wide.get(letter, 18)
    if row_count:
        taxonomy_column = get_column_letter(HEADERS.index("primary_taxonomy") + 1)
        recoverability_column = get_column_letter(
            HEADERS.index("text_recoverability") + 1
        )
        reviewed_column = get_column_letter(HEADERS.index("reviewed_status") + 1)
        taxonomy = DataValidation(
            type="list",
            formula1=f"=ValidationLists!$A$1:$A${len(ERROR_TAXONOMY)}",
        )
        recoverability = DataValidation(
            type="list",
            formula1=(
                f"=ValidationLists!$B$1:$B${len(RECOVERABILITY_LEVELS)}"
            ),
        )
        reviewed = DataValidation(type="list", formula1='"reviewed,pending"')
        for validation, column in (
            (taxonomy, taxonomy_column),
            (recoverability, recoverability_column),
            (reviewed, reviewed_column),
        ):
            sheet.add_data_validation(validation)
            validation.add(f"{column}2:{column}{row_count + 1}")


def create_rater_workbooks(
    records: Iterable[dict], output_dir: Path, seed: int, overwrite: bool = False
) -> list[Path]:
    """Create independently shuffled rater A and B workbooks."""

    source_records = list(records)
    output_dir.mkdir(parents=True, exist_ok=True)
    targets = []
    for offset, rater in enumerate(("A", "B")):
        target = output_dir / f"error_analysis_rater_{rater}.xlsx"
        if target.exists() and not overwrite:
            targets.append(target)
            continue
        rows = list(source_records)
        random.Random(seed + offset).shuffle(rows)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Error Analysis"
        sheet.append(HEADERS)
        for record in rows:
            sheet.append(_annotation_values(record))
        validation = workbook.create_sheet("ValidationLists")
        for row_index, label in enumerate(ERROR_TAXONOMY, start=1):
            validation.cell(row=row_index, column=1, value=label)
        for row_index, label in enumerate(RECOVERABILITY_LEVELS, start=1):
            validation.cell(row=row_index, column=2, value=label)
        validation.sheet_state = "hidden"
        instructions = workbook.create_sheet("Instructions")
        instructions.append(
            [
                "Annotate independently. Fill primary_taxonomy, optional comma-separated "
                "secondary_tags, text_recoverability, annotator, and set "
                "reviewed_status=reviewed. Do not inspect the other rater workbook."
            ]
        )
        instructions.append(
            [
                "raw_asr rows support recoverability analysis; correction_introduced rows "
                "support over-correction taxonomy; correction_residual rows describe "
                "changed but unresolved errors. Blank labels are invalid."
            ]
        )
        _style_annotation_sheet(sheet, len(rows))
        workbook.save(target)
        targets.append(target)
    return targets


def _sheet_rows(path: Path, sheet_name: str = "Error Analysis") -> list[dict]:
    sheet = load_workbook(path, read_only=True, data_only=True)[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "") for value in rows[0]]
    return [dict(zip(headers, values)) for values in rows[1:]]


def create_adjudication_workbook(output_dir: Path, overwrite: bool = False) -> Path:
    """Merge complete rater workbooks into a separate adjudication template."""

    target = output_dir / "error_analysis_adjudication.xlsx"
    if target.exists() and not overwrite:
        return target
    rater_a = {
        str(row["error_id"]): row
        for row in _sheet_rows(output_dir / "error_analysis_rater_A.xlsx")
    }
    rater_b = {
        str(row["error_id"]): row
        for row in _sheet_rows(output_dir / "error_analysis_rater_B.xlsx")
    }
    if set(rater_a) != set(rater_b) or not rater_a:
        raise ValueError("Rater workbooks must contain identical non-empty error IDs")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Adjudication"
    headers = [
        "error_id",
        "sample_id",
        "error_source",
        "reference",
        "raw_transcript",
        "corrected_transcript",
        "reference_span",
        "hypothesis_span",
        "rater_A_taxonomy",
        "rater_B_taxonomy",
        "final_primary_taxonomy",
        "rater_A_recoverability",
        "rater_B_recoverability",
        "final_text_recoverability",
        "agreement_status",
        "adjudicator",
        "adjudication_notes",
    ]
    sheet.append(headers)
    for error_id in sorted(rater_a):
        first, second = rater_a[error_id], rater_b[error_id]
        taxonomy_agreed = (
            first.get("primary_taxonomy")
            and first.get("primary_taxonomy") == second.get("primary_taxonomy")
        )
        recoverability_agreed = (
            first.get("text_recoverability")
            and first.get("text_recoverability")
            == second.get("text_recoverability")
        )
        agreed = bool(taxonomy_agreed and recoverability_agreed)
        sheet.append(
            [
                error_id,
                first.get("sample_id"),
                first.get("error_source"),
                first.get("reference"),
                first.get("raw_transcript"),
                first.get("corrected_transcript"),
                first.get("reference_span"),
                first.get("hypothesis_span"),
                first.get("primary_taxonomy"),
                second.get("primary_taxonomy"),
                first.get("primary_taxonomy") if taxonomy_agreed else "",
                first.get("text_recoverability"),
                second.get("text_recoverability"),
                first.get("text_recoverability") if recoverability_agreed else "",
                "agreed" if agreed else "needs_adjudication",
                "",
                "",
            ]
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)
    return target


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--config",
        default="dataset_benchmark/robustness_v2/configs/error_analysis_config.json",
    )
    parser.add_argument("--overwrite", action="store_true")
    parser.add_argument("--build-adjudication", action="store_true")
    args = parser.parse_args()
    config_path = resolve_path(args.config)
    config = _config(config_path)
    output_dir = resolve_path(config["annotation_dir"])
    if args.build_adjudication:
        rater_a = output_dir / "error_analysis_rater_A.xlsx"
        rater_b = output_dir / "error_analysis_rater_B.xlsx"
        target = output_dir / "error_analysis_adjudication.xlsx"
        existed_before = target.exists()
        result = create_adjudication_workbook(
            output_dir, overwrite=args.overwrite
        )
        atomic_write_json(
            resolve_path(config["stage_metadata"]["build_adjudication"]),
            build_stage_metadata(
                "build_error_analysis_adjudication",
                inputs={
                    "config": config_path,
                    "rater_A_workbook": rater_a,
                    "rater_B_workbook": rater_b,
                },
                outputs={"adjudication_workbook": result},
                details={
                    "execution_mode": (
                        "reuse_existing_output"
                        if existed_before and not args.overwrite
                        else "write_output"
                    ),
                    "output_written": not existed_before or args.overwrite,
                },
            ),
        )
        print(result)
        return
    manifest_path = resolve_path(config["manifest"])
    raw_path = resolve_path(config["raw_transcripts"])
    corrected_path = resolve_path(config["corrected_transcripts"])
    manifest = _read_manifest(manifest_path)
    raw = {
        int(row["audio_id"]): row
        for row in read_jsonl(raw_path)
    }
    corrected = {
        int(row["audio_id"]): row
        for row in read_jsonl(corrected_path)
    }
    records = build_annotation_records(manifest, raw, corrected)
    if not records:
        raise RuntimeError("No error spans were found for annotation")
    expected_targets = {
        "rater_A_workbook": output_dir / "error_analysis_rater_A.xlsx",
        "rater_B_workbook": output_dir / "error_analysis_rater_B.xlsx",
    }
    existed_before = {
        name: path.exists() for name, path in expected_targets.items()
    }
    targets = create_rater_workbooks(
        records,
        output_dir,
        seed=int(config["annotation_seed"]),
        overwrite=args.overwrite,
    )
    atomic_write_json(
        resolve_path(config["stage_metadata"]["prepare_annotation"]),
        build_stage_metadata(
            "prepare_error_analysis_annotation",
            inputs={
                "config": config_path,
                "manifest": manifest_path,
                "raw_transcripts": raw_path,
                "corrected_transcripts": corrected_path,
            },
            outputs=expected_targets,
            details={
                "annotation_rows": len(records),
                "execution_mode": (
                    "reuse_existing_outputs"
                    if all(existed_before.values()) and not args.overwrite
                    else "write_outputs"
                ),
                "outputs_written": {
                    name: not existed_before[name] or args.overwrite
                    for name in expected_targets
                },
            },
        ),
    )
    for target in targets:
        print(target)


if __name__ == "__main__":
    main()
