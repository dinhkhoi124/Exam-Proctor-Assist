"""Validate error annotations and evaluate recoverability/oracle bounds."""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path

from openpyxl import load_workbook

from dataset_benchmark.robustness_v2.error_analysis import (
    ERROR_TAXONOMY,
    RECOVERABILITY_LEVELS,
    build_stage_metadata,
    inter_rater_metrics,
    oracle_metrics,
)
from dataset_benchmark.scripts.common import atomic_write_json, read_jsonl, resolve_path
from dataset_benchmark.scripts.metrics import normalize_text


REVIEWED_VALUES = {"reviewed", "true", "yes", "complete"}


def worksheet_rows(path: Path, sheet_name: str) -> list[dict]:
    sheet = load_workbook(path, read_only=True, data_only=True)[sheet_name]
    values = list(sheet.iter_rows(values_only=True))
    if not values:
        return []
    headers = [str(value or "") for value in values[0]]
    return [dict(zip(headers, row)) for row in values[1:]]


def validate_rater_rows(rows: list[dict], rater_name: str) -> list[dict]:
    """Reject incomplete or invalid annotations; blanks are never labels."""

    if not rows:
        raise ValueError(f"{rater_name} workbook contains no annotation rows")
    validated = []
    for index, row in enumerate(rows, start=2):
        error_id = str(row.get("error_id") or "").strip()
        taxonomy = str(row.get("primary_taxonomy") or "").strip()
        secondary_tags = [
            value.strip()
            for value in str(row.get("secondary_tags") or "").split(",")
            if value.strip()
        ]
        recoverability = str(row.get("text_recoverability") or "").strip()
        annotator = str(row.get("annotator") or "").strip()
        reviewed = str(row.get("reviewed_status") or "").strip().casefold()
        problems = []
        if not error_id:
            problems.append("error_id is blank")
        if taxonomy not in ERROR_TAXONOMY:
            problems.append("primary_taxonomy is blank or invalid")
        invalid_secondary = [
            value for value in secondary_tags if value not in ERROR_TAXONOMY
        ]
        if invalid_secondary:
            problems.append(
                "secondary_tags contains invalid labels: "
                + ", ".join(invalid_secondary)
            )
        if recoverability not in RECOVERABILITY_LEVELS:
            problems.append("text_recoverability is blank or invalid")
        if not annotator:
            problems.append("annotator is blank")
        if reviewed not in REVIEWED_VALUES:
            problems.append("reviewed_status is not reviewed")
        if problems:
            raise ValueError(
                f"{rater_name} row {index} ({error_id or 'unknown'}): "
                + "; ".join(problems)
            )
        normalized = dict(row)
        normalized.update(
            {
                "error_id": error_id,
                "primary_taxonomy": taxonomy,
                "secondary_tags": secondary_tags,
                "text_recoverability": recoverability,
                "correction_resolved": bool(row.get("correction_resolved")),
                "transcript_changed": bool(row.get("transcript_changed")),
            }
        )
        validated.append(normalized)
    return validated


def final_annotations(
    rater_a: list[dict], rater_b: list[dict], adjudication_path: Path | None
) -> list[dict]:
    """Select agreed labels or require a complete adjudication decision."""

    first = {row["error_id"]: row for row in rater_a}
    second = {row["error_id"]: row for row in rater_b}
    if set(first) != set(second):
        raise ValueError("Rater workbooks contain different error IDs")
    adjudication = {}
    if adjudication_path and adjudication_path.exists():
        adjudication = {
            str(row.get("error_id") or ""): row
            for row in worksheet_rows(adjudication_path, "Adjudication")
        }
    output = []
    for error_id in sorted(first):
        row_a, row_b = first[error_id], second[error_id]
        taxonomy_agreed = (
            row_a["primary_taxonomy"] == row_b["primary_taxonomy"]
        )
        recoverability_agreed = (
            row_a["text_recoverability"] == row_b["text_recoverability"]
        )
        if taxonomy_agreed and recoverability_agreed:
            taxonomy = row_a["primary_taxonomy"]
            recoverability = row_a["text_recoverability"]
        else:
            decision = adjudication.get(error_id, {})
            status = str(decision.get("agreement_status") or "").casefold()
            taxonomy = str(decision.get("final_primary_taxonomy") or "")
            recoverability = str(decision.get("final_text_recoverability") or "")
            adjudicator = str(decision.get("adjudicator") or "").strip()
            if (
                status != "adjudicated"
                or taxonomy not in ERROR_TAXONOMY
                or recoverability not in RECOVERABILITY_LEVELS
                or not adjudicator
            ):
                raise ValueError(
                    f"Error {error_id} has unresolved rater disagreement; "
                    "complete the adjudication workbook"
                )
        final = dict(row_a)
        final["primary_taxonomy"] = taxonomy
        final["text_recoverability"] = recoverability
        output.append(final)
    return output


def evaluate(
    annotation_dir: Path, *, total_reference_words: int | None = None
) -> dict:
    rater_a = validate_rater_rows(
        worksheet_rows(annotation_dir / "error_analysis_rater_A.xlsx", "Error Analysis"),
        "rater A",
    )
    rater_b = validate_rater_rows(
        worksheet_rows(annotation_dir / "error_analysis_rater_B.xlsx", "Error Analysis"),
        "rater B",
    )
    agreement = inter_rater_metrics(rater_a, rater_b)
    adjudication_path = annotation_dir / "error_analysis_adjudication.xlsx"
    final = final_annotations(rater_a, rater_b, adjudication_path)
    return {
        "inter_rater_agreement": agreement,
        "oracle": oracle_metrics(
            final, total_reference_words=total_reference_words
        ),
    }


def reference_word_total(manifest_path: Path, raw_path: Path) -> int:
    """Count WER denominator words for every eligible sample with cached STT."""

    with manifest_path.open("r", encoding="utf-8-sig", newline="") as handle:
        manifest = list(csv.DictReader(handle))
    raw_ids = {int(row["audio_id"]) for row in read_jsonl(raw_path)}
    return sum(
        len(normalize_text(row["reference_transcript"]).split())
        for row in manifest
        if row.get("eligibility_status") == "eligible"
        and int(row["audio_id"]) in raw_ids
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--config",
        default="dataset_benchmark/robustness_v2/configs/error_analysis_config.json",
    )
    args = parser.parse_args()
    config_path = resolve_path(args.config)
    config = json.loads(config_path.read_text(encoding="utf-8"))
    annotation_dir = resolve_path(config["annotation_dir"])
    manifest_path = resolve_path(config["manifest"])
    raw_path = resolve_path(config["raw_transcripts"])
    rater_a_path = annotation_dir / "error_analysis_rater_A.xlsx"
    rater_b_path = annotation_dir / "error_analysis_rater_B.xlsx"
    adjudication_path = annotation_dir / "error_analysis_adjudication.xlsx"
    inputs = {
        "config": config_path,
        "manifest": manifest_path,
        "raw_transcripts": raw_path,
        "rater_A_workbook": rater_a_path,
        "rater_B_workbook": rater_b_path,
    }
    if adjudication_path.exists():
        inputs["adjudication_workbook"] = adjudication_path
    metadata_path = resolve_path(config["stage_metadata"]["evaluate_oracle"])
    try:
        result = evaluate(
            annotation_dir,
            total_reference_words=reference_word_total(
                manifest_path,
                raw_path,
            ),
        )
    except ValueError as exc:
        atomic_write_json(
            metadata_path,
            build_stage_metadata(
                "evaluate_oracle_recoverability",
                inputs=inputs,
                status="failed",
                details={"error": str(exc)},
            ),
        )
        parser.error(str(exc))
    output = resolve_path(config["oracle_output"])
    atomic_write_json(output, result)
    atomic_write_json(
        metadata_path,
        build_stage_metadata(
            "evaluate_oracle_recoverability",
            inputs=inputs,
            outputs={"oracle_report": output},
            details={"annotation_rows": result["inter_rater_agreement"]["items"]},
        ),
    )
    print(output)


if __name__ == "__main__":
    main()
